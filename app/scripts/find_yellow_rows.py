import openpyxl
from pathlib import Path

def find_yellow_rows():
    excel_path = Path("Input_Data/CARGOSOL LOGISTICS LTD CMA.xlsx")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\n==================================================")
        print(f"Sheet: {sheet_name}")
        print(f"==================================================")
        
        # Group by row
        row_fills = {}
        for r in range(1, sheet.max_row + 1):
            row_cols_filled = []
            for c in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=r, column=c)
                fill = cell.fill
                if fill and fill.fill_type and fill.fill_type != 'none':
                    color = fill.start_color
                    # Ignore default white fills (which might have rgb FFFFFFFF or similar, or theme 1)
                    if color:
                        # Check color attributes safely
                        color_type = color.type
                        rgb = None
                        try:
                            rgb = color.rgb
                        except Exception:
                            pass
                        indexed = color.indexed
                        theme = color.theme
                        tint = color.tint
                        
                        # We want to ignore typical white/gray fills
                        # Let's check if it's white/gray
                        if rgb == "00000000" or rgb == "FFFFFFFF" or rgb == "000000" or rgb == "FFFFFF":
                            continue
                        # If indexed is 64 or 65, it's typically system background/foreground
                        if indexed in [64, 65]:
                            continue
                            
                        row_cols_filled.append({
                            "col": c,
                            "val": cell.value,
                            "rgb": rgb,
                            "indexed": indexed,
                            "theme": theme,
                            "tint": tint
                        })
            if row_cols_filled:
                row_fills[r] = row_cols_filled
                
        # Now print the rows and their contents
        print(f"Found {len(row_fills)} colored rows:")
        for r, cols in row_fills.items():
            # Get values of this row
            row_vals = [sheet.cell(row=r, column=col_idx).value for col_idx in range(1, sheet.max_column + 1)]
            non_empty_vals = [f"Col{i+1}: {v}" for i, v in enumerate(row_vals) if v is not None]
            col_details = [f"Col{c['col']}(val={c['val']}, idx={c['indexed']}, theme={c['theme']}, rgb={c['rgb']})" for c in cols[:5]]
            print(f"Row {r}: {non_empty_vals[:4]}... -> Colored cols: {col_details}")

if __name__ == "__main__":
    find_yellow_rows()
