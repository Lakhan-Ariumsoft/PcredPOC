import openpyxl
from pathlib import Path

def dump_colored_rows():
    excel_path = Path("Input_Data/CARGOSOL LOGISTICS LTD CMA.xlsx")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    out_dir = Path("app/outputs/logs")
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "colored_rows.txt"
    
    with open(out_path, "w", encoding="utf-8") as f:
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            f.write(f"\n{'='*60}\n")
            f.write(f"Sheet: {sheet_name}\n")
            f.write(f"{'='*60}\n")
            
            for r in range(1, sheet.max_row + 1):
                colored_cells = []
                for c in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=r, column=c)
                    fill = cell.fill
                    if fill and fill.fill_type and fill.fill_type != 'none':
                        color = fill.start_color
                        if color:
                            rgb = None
                            try:
                                rgb = color.rgb
                            except Exception:
                                pass
                            
                            # Filter out default white/gray
                            if rgb in ["00000000", "FFFFFFFF", "000000", "FFFFFF"]:
                                continue
                            if color.indexed in [64, 65]:
                                continue
                                
                            colored_cells.append({
                                "col": c,
                                "val": cell.value,
                                "indexed": color.indexed,
                                "theme": color.theme,
                                "tint": color.tint,
                                "rgb": rgb
                            })
                            
                if colored_cells:
                    row_vals = [sheet.cell(row=r, column=col_idx).value for col_idx in range(1, sheet.max_column + 1)]
                    f.write(f"Row {r:3d}: {row_vals} | Colored: {colored_cells}\n")
                    
    print(f"Dumped colored rows details to {out_path}")

if __name__ == "__main__":
    dump_colored_rows()
