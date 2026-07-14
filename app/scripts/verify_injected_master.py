import openpyxl
from pathlib import Path

def verify_injected_cma():
    original_path = Path("Input_Data/CARGOSOL LOGISTICS LTD CMA.xlsx")
    wb = openpyxl.load_workbook(original_path, data_only=True)
    ws = wb["CMA"]
    
    def get_val(row, col):
        val = ws.cell(row=row, column=col).value
        try:
            return float(val) if val is not None else 0.0
        except (ValueError, TypeError):
            return 0.0

    print("=== Complete 3-Year CMA Sheet Balance Check (FY23, FY24, FY25) ===")
    
    years = {
        "FY23 (2022-23)": {"col": 8, "i_gen_res": 619.73, "i_surplus": 0.0},
        "FY24 (2023-24)": {"col": 9, "i_gen_res": 619.73, "i_surplus": 326.77}, # H138 + H146
        "FY25 (2024-25)": {"col": 10, "i_gen_res": 946.50, "i_surplus": -362.74} # I138 + I146
    }
    
    for name, info in years.items():
        col = info["col"]
        print(f"\n--- Checking {name} (Column {openpyxl.utils.get_column_letter(col)}) ---")
        
        # Load inputs directly from sheet
        sales = get_val(14, col)
        exp_sales = get_val(15, col)
        net_other_inc = get_val(17, col)
        
        freight = get_val(21, col)
        vehicle = get_val(22, col)
        dep = get_val(29, col)
        
        selling = get_val(46, col)
        
        # CC / TL / Other Interest detail rows 84, 85, 86
        if col == 10: # FY25
            cc_int = get_val(84, col)
            tl_int = get_val(85, col)
            other_int = get_val(86, col)
        else:
            cc_int = get_val(50, col)
            tl_int = get_val(51, col)
            other_int = get_val(52, col)
            
        tax = get_val(70, col)
        
        # Share Capital
        share_cap = get_val(137, col)
        if col in [9, 10]:
            share_cap = 1020.00 # formula =H137 or =I137
            
        gen_res = info["i_gen_res"] + info["i_surplus"]
        premium = get_val(144, col)
        
        # Liabilities inputs
        cc_loans = get_val(102, col)
        other_cc_loans = get_val(103, col)
        sundry_creditors = get_val(108, col)
        cust_advances = get_val(109, col)
        stat_dues = get_val(112, col)
        curr_maturities = get_val(114, col)
        
        if col == 8: # FY23
            other_cl = 0.88 + 39.24 + 4.47
            prov_others = 18.58
        elif col == 9: # FY24
            other_cl = 209.91 - cust_advances
            prov_others = 15.89
        else: # FY25
            other_cl = 7.69 + 4.31 + 49.56
            prov_others = 20.67
            
        term_loan = get_val(125, col)
        dir_loan = get_val(130, col)
        sec_dep_liab = get_val(131, col)
        unsec_loan = get_val(132, col)
        lt_prov = get_val(133, col)
        
        # Assets inputs
        cash_bal = get_val(158, col)
        fd_banks = get_val(160, col)
        receivables = get_val(161, col)
        if col == 10:
            receivables = 123.12 + 2677.59
        def_receivables = get_val(164, col)
        refund_govt = get_val(165, col)
        advances_supp = get_val(172, col)
        prepaid = get_val(176, col)
        deposits_emd = get_val(177, col)
        
        if col == 8:
            others_ca = 10.74
        elif col == 9:
            others_ca = 33.54
        else: # FY25
            others_ca = 145.97 - advances_supp + 19.00
            
        # Fixed Assets
        if col == 8:
            gross_block = 1394.25
            cwip = 1954.00
            dep_to_date = 826.04
        elif col == 9:
            gross_block = 3343.02
            cwip = 98.34
            dep_to_date = 1296.38
        else: # FY25
            gross_block = 3343.02 + 98.34 # formula =I182+I183
            cwip = 18.21 - 123.19         # formula =18.21-123.19
            dep_to_date = 1296.38 + dep - 65.64 # formula =I184+J29-65.64
            
        inv_others = get_val(188, col)
        def_rec_lt = get_val(190, col)
        
        if col == 8:
            sec_dep_lt = 22.78
        elif col == 9:
            sec_dep_lt = 25.30 + 10.28
        else: # FY25
            sec_dep_lt = 10.58 + 19.07
            
        dta = get_val(194, col)
        adv_tax = get_val(195, col)
        fd_lt = get_val(196, col)

        # Non-operating Items
        if col == 8:
            div_roy = 145.39
            other_inc = 15.63
        elif col == 9:
            div_roy = 133.77 + 7.47
            other_inc = 18.90 + 4.84 + 18.15
        else: # FY25
            div_roy = 0.04 + 5.78
            other_inc = 19.97 + 10.46 + 4.70 + 0.24 + 0.08 - 14.76
            
        # P&L Calculations
        gross_sales = sales + exp_sales
        total_income = gross_sales + net_other_inc
        cost_of_sales = freight + vehicle + dep
        gp = total_income - cost_of_sales
        
        if col == 8:
            admin = 1530.51
        elif col == 9:
            admin = 706.48 + 469.51 - selling
        else:
            admin = 716.42 + 417.11 - selling
            
        sub_exp = cost_of_sales + selling + admin
        op_profit = total_income - sub_exp
        total_int = cc_int + tl_int + other_int
        op_profit_after = op_profit - total_int
        
        net_non_op = round(div_roy + other_inc, 2)
        pbt = net_non_op + op_profit_after
        pat = pbt - tax
        cash_accruals = pat + dep
        
        # Liabilities Calculations
        sub_total_cl_a = cc_loans + other_cc_loans
        other_cl_subtotal = round(other_cl + prov_others, 2)
        sub_total_cl_b = sundry_creditors + cust_advances + stat_dues + curr_maturities + other_cl_subtotal
        total_cl = sub_total_cl_a + sub_total_cl_b
        
        other_tl_subtotal = round(dir_loan + sec_dep_liab + unsec_loan + lt_prov, 2)
        total_tl = term_loan + other_tl_subtotal
        outside_liab = total_cl + total_tl
        
        surplus = pat if total_income > 0 else 0.0
        net_worth = share_cap + gen_res + premium + surplus
        total_liab = outside_liab + net_worth
        
        # Assets Calculations
        other_ca_subtotal = round(prepaid + deposits_emd + others_ca, 2)
        total_ca = cash_bal + fd_banks + receivables + def_receivables + refund_govt + advances_supp + other_ca_subtotal
        net_block = gross_block + cwip - dep_to_date
        total_other_nca = inv_others + def_rec_lt + sec_dep_lt + dta + adv_tax + fd_lt
        total_assets = total_other_nca + net_block + total_ca
        
        diff = abs(total_assets - total_liab)
        
        print(f"  P&L: Total Income = {total_income:.2f} | COS = {cost_of_sales:.2f} | Interest = {total_int:.2f}")
        print(f"       PBT = {pbt:.2f} | Tax = {tax:.2f} | PAT = {pat:.2f} | Cash Accruals = {cash_accruals:.2f}")
        print(f"  Liabilities: Current = {total_cl:.2f} | Term = {total_tl:.2f} | Net Worth = {net_worth:.2f} (Cap={share_cap:.2f}, Res={gen_res:.2f}, Surp={surplus:.2f})")
        print(f"               TOTAL LIABILITIES = {total_liab:.2f}")
        print(f"  Assets     : Current = {total_ca:.2f} | Net Block = {net_block:.2f} | Other NCA = {total_other_nca:.2f}")
        print(f"               TOTAL ASSETS      = {total_assets:.2f}")
        print(f"  Difference: {diff:.4f} Lakhs")
        if diff < 0.15:
            print("  STATUS: Balanced! (OK)")
        else:
            print("  STATUS: MISMATCH! (ERROR)")

if __name__ == "__main__":
    verify_injected_cma()
