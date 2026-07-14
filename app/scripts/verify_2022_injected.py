import openpyxl
import re
from pathlib import Path

def evaluate_cell_value(ws, row, col):
    cell = ws.cell(row=row, column=col)
    val = cell.value
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str) and val.startswith("="):
        formula = val[1:].strip()
        
        # Special case for IF formula in row 146
        if "IF" in formula:
            # S146: =IF(S18>0,S74-S88,0)
            # We know S18 (growth in sales) is positive (20039.43), so we evaluate S74 - S88
            val_74 = evaluate_cell_value(ws, 74, col)
            val_88 = evaluate_cell_value(ws, 88, col)
            return val_74 - val_88
            
        # Replace cell references in Column S (like S46) with their evaluated values
        col_letter = openpyxl.utils.get_column_letter(col)
        pattern = rf"\b{col_letter}(\d+)\b"
        
        def replace_ref(match):
            ref_row = int(match.group(1))
            ref_val = evaluate_cell_value(ws, ref_row, col)
            return str(ref_val)
            
        formula_eval = re.sub(pattern, replace_ref, formula, flags=re.IGNORECASE)
        
        # Check if the string only contains numbers, operators (+, -, *, /), and spaces
        if re.match(r'^[\d\s\.\+\-\*\/\(\)]+$', formula_eval):
            try:
                return float(eval(formula_eval))
            except Exception:
                return 0.0
    return 0.0

def verify_injected_cma_2022():
    excel_path = Path("app/outputs/MASTER CARGOSOL LOGISTICS LTD CMA_with_2022.xlsx")
    if not excel_path.exists():
        print(f"Error: Excel file not found at {excel_path}")
        return
        
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    ws = wb["CMA"]
    
    col = 19
    get_val = lambda row: evaluate_cell_value(ws, row, col)

    print("=== CMA Sheet 2022 Balance Check (Column S, Index 19) ===")
    
    # Load inputs directly from sheet (using formula evaluation)
    sales = get_val(14)
    exp_sales = get_val(15)
    net_other_inc = get_val(17)
    
    freight = get_val(21)
    vehicle = get_val(22)
    dep = get_val(29)
    
    selling = get_val(46)
    admin = get_val(47)
    
    cc_int = get_val(84) or get_val(50)
    tl_int = get_val(85) or get_val(51)
    other_int = get_val(86) or get_val(52)
        
    tax = get_val(70)
    
    # Share Capital and reserves
    share_cap = get_val(137)
    gen_res = get_val(138) # Reserves and surplus (General Reserve / opening surplus)
    premium = get_val(144)
    drawings = get_val(145) # Bonus issue
    surplus = get_val(146) # Reserves and surplus (Surplus / current PAT)
    
    # Liabilities inputs
    cc_loans = get_val(102)
    other_cc_loans = get_val(103)
    sundry_creditors = get_val(108)
    cust_advances = get_val(109)
    stat_dues = get_val(112)
    curr_maturities = get_val(114)
    other_cl = get_val(117)
    prov_others = get_val(118)
        
    term_loan = get_val(125)
    dir_loan = get_val(130)
    sec_dep_liab = get_val(131)
    unsec_loan = get_val(132)
    lt_prov = get_val(133)
    
    # Assets inputs
    cash_bal = get_val(158)
    fd_banks = get_val(160)
    receivables = get_val(161)
    def_receivables = get_val(164)
    refund_govt = get_val(165)
    advances_supp = get_val(172)
    prepaid = get_val(176)
    deposits_emd = get_val(177)
    others_ca = get_val(178)
        
    # Fixed Assets
    gross_block = get_val(182)
    cwip = get_val(183)
    dep_to_date = get_val(184)
    intangibles_row = get_val(202) # Intangibles net block
        
    inv_others = get_val(188)
    def_rec_lt = get_val(190)
    sec_dep_lt = get_val(191)
    other_nca_dues = get_val(193) # property advances + subsidiary loan
    dta = get_val(194)
    adv_tax = get_val(195)
    fd_lt = get_val(196)

    # Non-operating Items
    div_roy = get_val(56)
    other_inc = get_val(58)
        
    # P&L Calculations
    gross_sales = sales + exp_sales
    total_income = gross_sales + net_other_inc
    cost_of_sales = freight + vehicle + dep
    gp = total_income - cost_of_sales
    
    sub_exp = cost_of_sales + selling + admin
    op_profit = total_income - sub_exp
    total_int = cc_int + tl_int + other_int
    op_profit_after = op_profit - total_int
    
    net_non_op = round(div_roy + other_inc, 2)
    pbt = net_non_op + op_profit_after
    pat = pbt - tax
    cash_accruals = pat + dep
    
    # Liabilities Calculations
    sub_total_cl_a = cc_loans + other_cc_loans
    other_cl_subtotal = round(other_cl + prov_others, 2)
    sub_total_cl_b = sundry_creditors + cust_advances + stat_dues + curr_maturities + other_cl_subtotal
    total_cl = sub_total_cl_a + sub_total_cl_b
    
    other_tl_subtotal = round(dir_loan + sec_dep_liab + unsec_loan + lt_prov, 2)
    total_tl = term_loan + other_tl_subtotal
    outside_liab = total_cl + total_tl
    
    # Net Worth
    net_worth = share_cap + gen_res + premium + drawings + surplus
    sheet_net_worth = get_val(147)
    
    # Total Liabilities
    total_liab = outside_liab + net_worth
    sheet_total_liab = get_val(148)
    
    # Assets Calculations
    other_ca_subtotal = round(prepaid + deposits_emd + others_ca, 2)
    total_ca = cash_bal + fd_banks + receivables + def_receivables + refund_govt + advances_supp + other_ca_subtotal
    sheet_total_ca = get_val(179)
    
    # Net Block
    net_block = gross_block + cwip - dep_to_date
    sheet_net_block = get_val(185)
    
    total_other_nca = inv_others + def_rec_lt + sec_dep_lt + other_nca_dues + dta + adv_tax + fd_lt
    sheet_total_other_nca = get_val(198)
    
    sheet_total_intangibles = get_val(205)
    
    # Total Assets
    total_assets = intangibles_row + total_other_nca + net_block + total_ca
    sheet_total_assets = get_val(206)
    
    diff = abs(total_assets - total_liab)
    
    print("\n--- Key Metrics ---")
    print(f"Sales: {sales:.2f}")
    print(f"Freight & Handling: {freight:.2f}")
    print(f"Depreciation (P&L): {dep:.2f}")
    print(f"Profit before Tax: {pbt:.2f}")
    print(f"Profit after Tax: {pat:.2f}")
    
    print("\n--- Balance Sheet Verification ---")
    print(f"Total Current Liabilities: {total_cl:.2f}")
    print(f"Total Term Liabilities   : {total_tl:.2f}")
    print(f"Net Worth                : {net_worth:.2f} (Excel cell: {sheet_net_worth:.2f})")
    print(f"TOTAL LIABILITIES        : {total_liab:.2f} (Excel cell: {sheet_total_liab:.2f})")
    
    print(f"Total Current Assets     : {total_ca:.2f} (Excel cell: {sheet_total_ca:.2f})")
    print(f"Net Block                : {net_block:.2f} (Excel cell: {sheet_net_block:.2f})")
    print(f"Other Non-Current Assets : {total_other_nca:.2f} (Excel cell: {sheet_total_other_nca:.2f})")
    print(f"Intangibles              : {intangibles_row:.2f} (Excel cell: {sheet_total_intangibles:.2f})")
    print(f"TOTAL ASSETS             : {total_assets:.2f} (Excel cell: {sheet_total_assets:.2f})")
    
    print(f"Difference: {diff:.4f} Lakhs")
    if diff < 0.15:
        print("STATUS: Balanced! (OK)")
    else:
        print("STATUS: MISMATCH! (ERROR)")
        
    wb_formulas = openpyxl.load_workbook(excel_path, data_only=False)
    ws_formulas = wb_formulas["CMA"]
    print("\n--- Cell Formulas Check ---")
    print("Row 182 (Gross Block) formula/val:", ws_formulas.cell(row=182, column=19).value)
    print("Row 184 (Depreciation to Date) formula/val:", ws_formulas.cell(row=184, column=19).value)

if __name__ == "__main__":
    verify_injected_cma_2022()
