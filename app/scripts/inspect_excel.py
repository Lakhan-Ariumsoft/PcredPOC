import openpyxl
from pathlib import Path

def inspect_excel():
    excel_path = Path("Input_Data/CARGOSOL LOGISTICS LTD CMA.xlsx")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb['CMA']
    count = 0
    print("Listing cells with non-white, non-empty fills in sheet CMA:")
    for r in range(1, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=r, column=c)
            fill = cell.fill
            if fill and fill.fill_type and fill.fill_type != 'none':
                color = fill.start_color
                rgb = color.rgb if color else None
                # Ignore white fills
                if rgb and str(rgb).upper() in ["00000000", "FFFFFFFF", "000000", "FFFFFF"]:
                    continue
                # Print details of the fill
                count += 1
                theme = color.theme if color else None
                indexed = color.indexed if color else None
                tint = color.tint if color else None
                print(f"Cell {r},{c} val={cell.value}: fill_type={fill.fill_type}, rgb={rgb}, theme={theme}, indexed={indexed}, tint={tint}")
                if count >= 50:
                    return

if __name__ == "__main__":
    inspect_excel()
