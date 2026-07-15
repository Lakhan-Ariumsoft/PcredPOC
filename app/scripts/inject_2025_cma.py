import openpyxl
import shutil
from pathlib import Path

def inject_cma_data():
    original_path = Path("Input_Data/CARGOSOL LOGISTICS LTD CMA.xlsx")
    backup_path = Path("app/outputs/CARGOSOL LOGISTICS LTD CMA_backup.xlsx")
    
    # 1. Create a backup
    backup_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(original_path, backup_path)
    print(f"Backup created at: {backup_path}")
    
    # 2. Load the workbook (preserving formulas)
    wb = openpyxl.load_workbook(original_path, data_only=False)
    
    # 3. Update 'CMA' sheet
    ws_cma = wb["CMA"]
    print("Updating 'CMA' sheet...")
    
    # We specify column J (col 10) overrides
    cma_overrides = {
        12: 12,                                      # Operating months
        14: 13200.24,                                # i. Domestic Sale (Note 21)
        15: None,                                    # ii.Export Sale
        17: None,                                    # Net Other Income
        21: 11476.01,                                # a. Freight & Handling Expenses (Note 23)
        22: 20.35,                                   # b. Vehicle Running expenses (Note 23)
        23: None,                                    # c. Stores & Spares (Imported)
        24: None,                                    # d. Stores & Spares (Indigenous)
        25: None,                                    # Power & Fuel
        26: None,                                    # Direct Labour
        27: None,                                    # Repairs and maintainance
        28: None,                                    # Other Mfg. Expenses
        29: 262.54,                                  # Depreciation (P&L)
        46: 48.88,                                   # Selling Expenses (Note 26, Business Promotion)
        47: "=716.42+417.11-J46",                    # Administrative Expenses (P&L Admin formula)
        50: "=J84",                                  # a. Interest on CC
        51: "=J85",                                  # b.Interest on TL
        52: "=J86",                                  # c.Other interests
        56: "=0.04+5.78",                            # Interest/Dividend/Royalties etc. (Note 22)
        58: "=19.97+10.46+4.7+0.24+0.08-14.76",      # Other Income (Note 22 details)
        70: 9.81,                                    # Provision for Taxes (Deferred Tax, P&L)
        85: 246.21,                                  # Interest on TL (Note 25, Bank Interest)
        86: 65.90,                                   # Other interests (Note 25, Other Borrowing Costs)
        102: 1266.13,                                # Short Term loans Applicant Bank (Note 7)
        103: 83.09,                                  # Short Term loans Other Banks (Note 7)
        108: 847.32,                                 # Sundry Creditors (Trade, Balance Sheet)
        109: 85.39,                                  # Advance Payment from Customers (Note 9)
        112: 109.11,                                 # Other Statutory Liab (Note 9, Statutory Dues)
        114: 281.67,                                 # Installments of term Loan (Note 7, Current maturities)
        117: "=7.69+4.31+49.56",                      # Other current Liabilities (Note 9)
        118: 20.67,                                  # Provision for Others (Note 10, Provisions)
        125: 569.79,                                 # Term Loan from Bank (Note 4, Secured Loans Total)
        130: 166.23,                                 # Unsecured Loans from Directors (Note 4)
        131: 3.00,                                   # Security Deposits (Note 5)
        132: 144.80,                                 # Unsecured Loans (Note 4, Unsecured Loans Total)
        133: 55.66,                                  # Long Term Provisions (Note 6)
        144: 388.60,                                 # Share Premium (Note 3)
        158: 52.10,                                  # Cash & Bank Balances (Note 17)
        160: 42.56,                                  # Fixed Deposits with Banks (Note 18)
        161: "=123.12+2677.59",                      # Domestic Receivables (Note 16 current)
        164: 2.70,                                   # Deferred receivables (Note 16 6m-1y)
        165: 1.58,                                   # Refund Dues from Govt (Note 20)
        172: 140.95,                                 # Advances to Suppliers (Note 19)
        176: 62.43,                                  # Prepaid expenses (Note 20)
        177: 8.93,                                   # Deposits-EMD & Fixed Deposits (Note 20)
        178: "=145.97-J172+19",                      # Others (Current Assets formula)
        182: "=I182+I183",                           # Gross Block (original formula)
        183: "=18.21-123.19",                        # CWIP (original formula)
        184: "=I184+J29-65.64",                      # Depreciation to Date (original formula)
        188: 48.09,                                  # Investment in Others (Note 12)
        190: 241.34,                                 # Deferred Receivables LT (Note 16 disputed)
        191: "=10.58+19.07",                          # Security Deposits LT (Note 15)
        194: 156.95,                                 # Deferred Tax Asset (net) (Note 13)
        195: 193.20,                                 # Advance Tax/TDS (Note 14)
        196: 50.88                                   # Fixed Deposits LT (Note 15)
    }
    
    for row, val in cma_overrides.items():
        cell = ws_cma.cell(row=row, column=10)
        print(f"  Row {row:03d}: Writing '{val}' (previous: '{cell.value}')")
        cell.value = val
        
    # 4. Update 'DEPRECIATION' sheet (01.04.2025 opening values)
    ws_dep = wb["DEPRECIATION"]
    print("\nUpdating 'DEPRECIATION' sheet...")
    
    # We specify column C (Gross block on 01.04.2025) and column F (Depreciation on 01.04.2025) overrides
    # Mapping: Row -> (col_C_val, col_F_val)
    dep_overrides = {
        4: (5.20, None),          # Land
        5: (1195.36, 190.06),      # Office Premises
        6: (122.43, 82.87),        # Furniture & Fixtures
        7: (1189.08, 613.12),      # Container
        8: (89.70, 74.77),         # Office Equipment
        9: (582.47, 399.35),       # Vehicles
        10: (114.70, 104.06),      # Computer
        12: (30.64, 29.05)         # Software
    }
    
    for row, (c_val, f_val) in dep_overrides.items():
        cell_c = ws_dep.cell(row=row, column=3)
        print(f"  Row {row:02d} Col C: Writing '{c_val}' (previous: '{cell_c.value}')")
        cell_c.value = c_val
        
        if f_val is not None:
            cell_f = ws_dep.cell(row=row, column=6)
            print(f"  Row {row:02d} Col F: Writing '{f_val}' (previous: '{cell_f.value}')")
            cell_f.value = f_val
            
    # Save the updated workbook back
    wb.save(original_path)
    print(f"\nSuccessfully updated master Excel sheet at {original_path}")

if __name__ == "__main__":
    inject_cma_data()
