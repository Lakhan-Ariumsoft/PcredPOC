import openpyxl
from pathlib import Path

def find_yellow_rows_only():
    excel_path = Path("Input_Data/CARGOSOL LOGISTICS LTD CMA.xlsx")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    yellow_rows = []
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for r in range(1, sheet.max_row + 1):
            is_yellow_row = False
            yellow_cells = []
            for c in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=r, column=c)
                fill = cell.fill
                if fill and fill.fill_type and fill.fill_type != 'none':
                    color = fill.start_color
                    if color:
                        try:
                            rgb = color.rgb
                            if rgb == "FFFFFF00" or rgb == "FFFF00":
                                is_yellow_row = True
                                yellow_cells.append(c)
                        except Exception:
                            pass
            if is_yellow_row:
                row_vals = [sheet.cell(row=r, column=col_idx).value for col_idx in range(1, sheet.max_column + 1)]
                # Filter out None from values for display
                non_empty = [f"Col{i+1}: {v}" for i, v in enumerate(row_vals) if v is not None]
                yellow_rows.append((sheet_name, r, row_vals, yellow_cells, non_empty))
                
    print(f"Total yellow rows found: {len(yellow_rows)}")
    for s_name, r_idx, vals, cols, non_empty in yellow_rows:
        print(f"\nSheet: {s_name} | Row {r_idx} | Yellow Columns: {cols}")
        print(f"Values: {non_empty}")

if __name__ == "__main__":
    find_yellow_rows_only()
