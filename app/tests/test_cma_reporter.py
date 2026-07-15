import pytest
import io
from openpyxl import load_workbook
from app.services.cma_reporter_service import generate_cma_excel

def test_generate_cma_excel():
    mock_data = {
        "company_name": "Acme Corp",
        "company_slug": "acme-corp",
        "financial_years": ["2022-23", "2023-24", "2024-25"],
        "cma_data": {
            "sales": {
                "label": "Sales",
                "fields": {
                    "Net Sales": {
                        "2022-23": {"value": 1000.0, "confidence": 0.9},
                        "2023-24": {"value": 1100.0, "confidence": 0.95},
                        "2024-25": {"value": 1200.0, "confidence": 0.95}
                    }
                }
            },
            "cost_of_sales": {
                "label": "Cost of Sales",
                "fields": {
                    "Depreciation": {
                        "2022-23": {"value": 50.0, "confidence": 0.9},
                        "2023-24": {"value": 55.0, "confidence": 0.9},
                        "2024-25": {"value": 60.0, "confidence": 0.9}
                    }
                }
            }
        }
    }
    
    excel_bytes = generate_cma_excel(mock_data)
    
    assert isinstance(excel_bytes, bytes)
    # Excel files are zip archives starting with the standard PK header
    assert excel_bytes.startswith(b"PK\x03\x04")
    
    # Load with openpyxl to inspect sheets
    wb = load_workbook(io.BytesIO(excel_bytes))
    assert "CMA" in wb.sheetnames
    
    ws = wb["CMA"]
    # Check that title cell contains Acme Corp
    title_val = ws.cell(row=1, column=1).value
    assert "Acme Corp" in title_val


def _value_columns():
    from app.services.cma_reporter_service import ALL
    return ALL


def test_generate_cma_excel_never_injects_hardcoded_template_numbers():
    """
    Regression guard: cma_reporter_service used to hardcode a batch of
    Cargosol-Logistics-specific figures (Share Capital projection 1020.0,
    Gross Block projection 3336.38, "Others"/"Share Premium" fixed at
    388.6 for every single year including audited ones, etc.) as fallback
    values applied to every company's report regardless of that company's
    real data. A company with none of these fields extracted must not see
    any of those numbers appear anywhere in its generated report.
    """
    mock_data = {
        "company_name": "No Extra Data Ltd",
        "company_slug": "no-extra-data-ltd",
        "financial_years": ["2022-23", "2023-24", "2024-25"],
        "cma_data": {
            "sales": {
                "label": "Sales",
                "fields": {
                    "Net Sales": {
                        "2022-23": {"value": 1000.0, "confidence": 0.9},
                        "2023-24": {"value": 1100.0, "confidence": 0.95},
                        "2024-25": {"value": 1200.0, "confidence": 0.95},
                    }
                },
            },
        },
    }
    excel_bytes = generate_cma_excel(mock_data)
    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb["CMA"]

    formerly_hardcoded_values = {
        256.5, 2350.0, 500, 169.23, 166.23, 3.0, 1020.0, 388.6,
        42.56, 400, 50.0, 11.0, 3336.38, 48.09, 29.65, 50.88, 53.0,
    }
    seen_values = set()
    for r in range(1, ws.max_row + 1):
        for c in _value_columns():
            v = ws.cell(row=r, column=c).value
            if isinstance(v, (int, float)):
                seen_values.add(v)

    leaked = formerly_hardcoded_values & seen_values
    assert not leaked, f"Hardcoded template values leaked into output: {leaked}"


def test_generate_cma_excel_removes_fully_null_rows():
    mock_data = {
        "company_name": "Sparse Corp",
        "company_slug": "sparse-corp",
        "financial_years": ["2022-23", "2023-24", "2024-25"],
        "cma_data": {
            "sales": {
                "label": "Sales",
                "fields": {
                    "Net Sales": {
                        "2022-23": {"value": 1000.0, "confidence": 0.9},
                        "2023-24": {"value": 1100.0, "confidence": 0.95},
                        "2024-25": {"value": 1200.0, "confidence": 0.95},
                    }
                },
            },
        },
    }
    excel_bytes = generate_cma_excel(mock_data)
    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb["CMA"]

    value_cols = _value_columns()
    for r in range(1, ws.max_row + 1):
        label = ws.cell(row=r, column=3).value
        if label is None:
            continue  # section header / merged row — not a data row
        values = [ws.cell(row=r, column=c).value for c in value_cols]
        assert not all(v in (None, "") for v in values), (
            f"row {r} ('{label}') should have been removed — all value columns empty"
        )


def test_generate_cma_excel_keeps_populated_rows():
    mock_data = {
        "company_name": "Sparse Corp",
        "company_slug": "sparse-corp",
        "financial_years": ["2022-23", "2023-24", "2024-25"],
        "cma_data": {
            "sales": {
                "label": "Sales",
                "fields": {
                    "Net Sales": {
                        "2022-23": {"value": 1000.0, "confidence": 0.9},
                        "2023-24": {"value": 1100.0, "confidence": 0.95},
                        "2024-25": {"value": 1200.0, "confidence": 0.95},
                    }
                },
            },
        },
    }
    excel_bytes = generate_cma_excel(mock_data)
    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb["CMA"]

    labels = [ws.cell(row=r, column=3).value for r in range(1, ws.max_row + 1)]
    assert any(lbl and "Domestic Sale" in str(lbl) for lbl in labels)


def test_generate_cma_excel_appends_unmapped_items():
    mock_data = {
        "company_name": "Acme Corp",
        "company_slug": "acme-corp",
        "financial_years": ["2022-23"],
        "cma_data": {},
        "unmapped_items": [
            {
                "label": "Some Custom Line Item",
                "page": 7,
                "current_value": 123.45,
                "previous_value": 99.0,
                "source_filename": "test.pdf",
            }
        ],
    }
    excel_bytes = generate_cma_excel(mock_data)
    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb["CMA"]

    labels = [ws.cell(row=r, column=3).value for r in range(1, ws.max_row + 1)]
    assert any(lbl and "Some Custom Line Item" in str(lbl) for lbl in labels)


def test_generate_cma_excel_no_unmapped_items_adds_nothing():
    mock_data = {
        "company_name": "Acme Corp",
        "company_slug": "acme-corp",
        "financial_years": ["2022-23"],
        "cma_data": {},
        "unmapped_items": [],
    }
    # Should not raise, and should not add an "ADDITIONAL ITEMS" section.
    excel_bytes = generate_cma_excel(mock_data)
    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb["CMA"]
    values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert not any(v and "ADDITIONAL ITEMS" in str(v) for v in values)
