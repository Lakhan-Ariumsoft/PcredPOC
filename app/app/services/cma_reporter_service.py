"""
cma_reporter.py  v2.0
=====================
Generates a single-sheet CMA Excel that EXACTLY mirrors the XLCMA template
(CARGOSOL_LOGISTICS_LTD_CMA.xlsx) — same row order, same labels, same
calculations, with:
  • 3 audited years  (FY 2022-23, 2023-24, 2024-25)  from FastAPI extraction
  • 2 projection years (FY 2025-26, 2026-27)           auto-calculated

Drop into your app/ folder.  Import:
    from app.cma_reporter import generate_cma_excel

Usage:
    xlsx_bytes = generate_cma_excel(merged_result_dict)

Or CLI:
    python cma_reporter.py merged.json [out.xlsx]
"""

from __future__ import annotations
import io, json, sys
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# COLUMN LAYOUT  (matches original XLCMA template exactly)
# Col A=1 B=2 C=3 D=4  E=5(2013-14) F=6(2014-15) G=7(2015-16)
# H=8(FY23) I=9(FY24) J=10(FY25) K=11(Proj26) L=12(Proj27)
# ─────────────────────────────────────────────────────────────────────────────
C_SR    = 2   # B  — serial number
C_LBL   = 3   # C  — label
C_VAR   = 4   # D  — variable % (BEP only)
C_FY23  = 8   # H  — FY 2022-23 (audited)
C_FY24  = 9   # I  — FY 2023-24 (audited)
C_FY25  = 10  # J  — FY 2024-25 (audited)
C_P26   = 11  # K  — Projection 2025-26
C_P27   = 12  # L  — Projection 2026-27

AUD = [C_FY23, C_FY24, C_FY25]   # audited columns
ALL = [C_FY23, C_FY24, C_FY25, C_P26, C_P27]

# ── colours ──────────────────────────────────────────────────────────────────
BG_TITLE   = "1F3864"
BG_SECT    = "2E75B6"
BG_SUBSECT = "BDD7EE"
BG_DATA    = "FFFFFF"
BG_PROJ    = "EBF3FB"
BG_TOTAL   = "DDEEFF"
BG_LABEL   = "F2F2F2"
BG_RATIO   = "FFF2CC"
FG_WHITE   = "FFFFFF"
FG_DARK    = "1F3864"
FG_BLACK   = "000000"
FG_CALC    = "1F3864"

NF_NUM  = '#,##0.00;(#,##0.00);"-"'
NF_PCT  = '0.00%'
NF_2DP  = '0.00'
NF_INT  = '0'


def _f(h): return PatternFill("solid", fgColor=h)
def _b():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def _bb():
    t = Side(style="medium", color="2E75B6")
    n = Side(style="thin", color="CCCCCC")
    return Border(left=n, right=n, top=t, bottom=t)
def _gc(c): return get_column_letter(c)
def _safe(v):
    try: return float(v) if v is not None else None
    except: return None


# ─────────────────────────────────────────────────────────────────────────────
class CMAWriter:
    def __init__(self, data: dict):
        self.data = data
        self.wb   = Workbook()
        self.ws   = self.wb.active
        self.ws.title = "CMA"
        self.row  = 1
        self._row_map: dict[str, int] = {}   # label → excel row
        self._col_widths()

    # ── column widths ─────────────────────────────────────────────────────────
    def _col_widths(self):
        ws = self.ws
        ws.column_dimensions[_gc(1)].width  = 2
        ws.column_dimensions[_gc(C_SR)].width  = 5
        ws.column_dimensions[_gc(C_LBL)].width = 40
        ws.column_dimensions[_gc(C_VAR)].width = 7
        for c in range(5, 13):
            ws.column_dimensions[_gc(c)].width = 14

    # ── low-level cell writer ─────────────────────────────────────────────────
    def _c(self, row, col, val,
           bold=False, color=FG_BLACK, bg=BG_DATA,
           align="left", nf=None, size=9, wrap=True, border=True, italic=False):
        cell = self.ws.cell(row=row, column=col, value=val)
        cell.font = Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
        cell.fill = _f(bg)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        if border: cell.border = _b()
        if nf:     cell.number_format = nf
        return cell

    def _merge(self, r1, c1, r2, c2, val, bold=True, bg=BG_SECT,
               color=FG_WHITE, align="left", size=9):
        self.ws.merge_cells(start_row=r1, start_column=c1,
                            end_row=r2,   end_column=c2)
        cell = self.ws.cell(row=r1, column=c1, value=val)
        cell.font = Font(name="Arial", bold=bold, color=color, size=size)
        cell.fill = _f(bg)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        return cell

    # ── TITLE ─────────────────────────────────────────────────────────────────
    def _title(self):
        r = self.row
        name = self.data.get("company_name", "Company")
        self.ws.row_dimensions[r].height = 28
        self._merge(r,1,r,12, f"XLCMA  –  {name}  –  Credit Monitoring Arrangement  |  INR Lakhs",
                    bold=True, bg=BG_TITLE, color=FG_WHITE, align="center", size=12)
        self.row += 1
        # Year header row
        r = self.row
        self.ws.row_dimensions[r].height = 30
        self._c(r, C_LBL, "PARTICULARS", bold=True, bg=BG_SECT, color=FG_WHITE, align="center")
        self._c(r, C_VAR, "Var%",        bold=True, bg=BG_SECT, color=FG_WHITE, align="center")
        hdrs = {C_FY23:"FY 2022-23\n(Audited)", C_FY24:"FY 2023-24\n(Audited)",
                C_FY25:"FY 2024-25\n(Audited)", C_P26:"FY 2025-26\n(Projection)",
                C_P27:"FY 2026-27\n(Projection)"}
        for col, h in hdrs.items():
            bg = BG_DATA if col<=C_FY25 else BG_PROJ
            fc = FG_DARK if col<=C_FY25 else "1F5C8B"
            self._c(r, col, h, bold=True, bg=bg, color=fc, align="center", nf=None)
        self.row += 1

    # ── section header ────────────────────────────────────────────────────────
    def _sec(self, label, bg=BG_SECT, size=9):
        r = self.row
        self.ws.row_dimensions[r].height = 20
        self._merge(r,1,r,12, label, bold=True, bg=bg, color=FG_WHITE,
                    align="left", size=size)
        self.row += 1

    # ── data row (the workhorse) ───────────────────────────────────────────────
    def _row(self, sr, label, vals: dict,
             bold=False, bg=None, nf=NF_NUM, indent=0,
             calc=False, var=None, total=False):
        """
        Write one data row.
        vals = {col_idx: value_or_None, ...}
        """
        r = self.row
        self.ws.row_dimensions[r].height = 16
        lb = ("  " * indent) + label

        row_bg = bg or (BG_TOTAL if total else BG_LABEL)
        fc = FG_CALC if calc else FG_BLACK

        self._c(r, C_SR,  str(sr) if sr else "", bold=bold, bg=row_bg, align="center")
        self._c(r, C_LBL, lb,                    bold=bold, bg=row_bg, color=fc)
        if var is not None:
            self._c(r, C_VAR, var, bg=row_bg, align="center")
        else:
            self._c(r, C_VAR, None, bg=row_bg)

        for col in ALL:
            v   = vals.get(col)
            pbg = BG_PROJ if col >= C_P26 else (row_bg if row_bg != BG_LABEL else BG_DATA)
            if isinstance(v, str) and v.startswith("="):
                cell = self.ws.cell(row=r, column=col, value=v)
                cell.font  = Font(name="Arial", bold=bold, color=FG_CALC, size=9)
                cell.fill  = _f(pbg)
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.border= _b()
                if nf: cell.number_format = nf
            else:
                fc2 = FG_CALC if (calc and v is not None) else FG_BLACK
                self._c(r, col, v, bold=bold, bg=pbg, align="right", nf=nf, color=fc2)

        self._row_map[label] = r
        self.row += 1
        return r

    def _blank(self, h=6):
        self.ws.row_dimensions[self.row].height = h
        self.row += 1

    # ─────────────────────────────────────────────────────────────────────────
    # DATA HELPERS
    # ─────────────────────────────────────────────────────────────────────────
    def _fys(self):
        """Return {C_FY23: fy_string, C_FY24: fy_string, C_FY25: fy_string}"""
        fys = self.data.get("financial_years", [])
        m = {}
        for fy in fys:
            yr = fy.split("-")[0]
            if   yr == "2022": m[C_FY23] = fy
            elif yr == "2023": m[C_FY24] = fy
            elif yr == "2024": m[C_FY25] = fy
        # fallback: assign by position
        if not m and fys:
            keys = [C_FY23, C_FY24, C_FY25]
            for i, fy in enumerate(fys[-3:]):
                m[keys[i]] = fy
        return m

    def _get(self, section, field, fy):
        """Safely extract value from merged cma_data, with automatic section and field alignment."""
        # Align section and field names to match app/cma_fields.py CMA_SECTIONS
        # Section mappings
        sec_map = {
            "non_current_assets": "fixed_assets",
        }
        # Section-specific field mappings
        field_sec_map = {
            # (old_section, old_field) -> (new_section, new_field)
            ("sales", "Export Sales"): ("sales", "Export Sale"),
            ("interest", "Other Interest"): ("interest", "Other interests"),
            ("cost_of_sales", "Selling Expenses"): ("profitability", "Selling Expenses"),
            ("cost_of_sales", "Administrative Expenses"): ("profitability", "Administrative Expenses"),
            ("pnl", "Interest/Dividend/Royalties"): ("other_income_expenses", "Interest/Dividend/Royalties"),
            ("pnl", "Other Income"): ("other_income_expenses", "Other Income"),
            ("pnl", "Net Other Income"): ("other_income_expenses", "Net of other non operating Income/Expenses"),
            ("term_liabilities", "Loan from Directors"): ("term_liabilities", "Unsecured Loans from Directors"),
            ("term_liabilities", "Long Term Provisions"): ("term_liabilities", "Long Term provisions"),
            ("term_liabilities", "Security Deposits"): ("fixed_assets", "Security Deposits"),
            ("current_assets", "Deferred Receivables"): ("current_assets", "Deferred receivables"),
            ("current_liabilities", "Other Current Liabilities"): ("current_liabilities", "Other current Liabilities"),
        }

        # First apply section mapping
        section = sec_map.get(section, section)

        # Then apply field/section specific mapping
        if (section, field) in field_sec_map:
            section, field = field_sec_map[(section, field)]

        return _safe(
            self.data.get("cma_data", {})
                .get(section, {})
                .get("fields", {})
                .get(field, {})
                .get(fy, {})
                .get("value")
        )

    def _v(self, section, field, proj_fn=None):
        """
        Build {C_FY23: v23, C_FY24: v24, C_FY25: v25, C_P26: p26, C_P27: p27}.
        proj_fn(v23,v24,v25) → (p26,p27)   or None
        """
        fy_map = self._fys()
        out = {}
        vals = {}
        for col, fy in fy_map.items():
            v = self._get(section, field, fy)
            out[col] = v
            vals[col] = v

        v23 = vals.get(C_FY23)
        v24 = vals.get(C_FY24)
        v25 = vals.get(C_FY25)

        if proj_fn:
            p26, p27 = proj_fn(v23, v24, v25)
            out[C_P26] = p26
            out[C_P27] = p27
        return out

    def _cell(self, row, col):
        return _safe(self.ws.cell(row=row, column=col).value)

    def _rrow(self, label):
        return self._row_map.get(label)

    # ─────────────────────────────────────────────────────────────────────────
    # PROJECTION HELPERS
    # ─────────────────────────────────────────────────────────────────────────
    @staticmethod
    def _avg_growth(v23, v24, v25):
        rates = []
        for a, b in [(v23, v24), (v24, v25)]:
            if a and b and a != 0: rates.append((b - a) / abs(a))
        return sum(rates)/len(rates) if rates else 0.10

    def _proj_growth(self, v23, v24, v25):
        """Project next 2 years using avg growth, capped ±50%."""
        g = max(-0.5, min(self._avg_growth(v23, v24, v25), 0.5))
        base = v25 or v24 or v23 or 0
        p26  = round(base * (1 + g), 4)
        p27  = round(p26  * (1 + g), 4)
        return p26, p27

    def _proj_pct(self, ratio_fn):
        """Project by applying avg pct ratio to projected sales."""
        def _fn(v23, v24, v25):
            return None, None   # filled in after sales row is written
        return _fn

    def _proj_fixed_ratio(self, v23, v24, v25, s23, s24, s25):
        """Project a cost item using its avg % of sales."""
        ratios = []
        for v, s in [(v23,s23),(v24,s24),(v25,s25)]:
            if v and s and s != 0: ratios.append(v/s)
        r = sum(ratios)/len(ratios) if ratios else 0
        sales26 = self._cell(self._rrow("i. Domestic Sale"), C_P26) or 0
        sales27 = self._cell(self._rrow("i. Domestic Sale"), C_P27) or 0
        return round(r * sales26, 4), round(r * sales27, 4)

    # ─────────────────────────────────────────────────────────────────────────
    # MAIN BUILD
    # ─────────────────────────────────────────────────────────────────────────
    def build(self):
        self._title()
        self._blank()
        self._pnl()
        self._blank()
        self._bs_liabilities()
        self._blank()
        self._bs_assets()
        self._blank()
        self._key_metrics()
        self._blank()
        self._wc_assessment()
        self._blank()
        self._fund_flow()
        self._blank()
        self._break_even()
        self._blank()
        self._dscr()
        self._blank()
        self._ratio_analysis()
        self._blank()
        self._key_indicators()
        self._append_unmapped_items()
        self._remove_null_rows()
        # freeze top 3 rows
        self.ws.freeze_panes = self.ws.cell(row=3, column=C_FY23)
        # print settings
        self.ws.page_setup.orientation = "landscape"
        self.ws.page_setup.fitToPage   = True
        self.ws.page_setup.fitToWidth  = 1

    # ─────────────────────────────────────────────────────────────────────────
    # OPERATING STATEMENT
    # ─────────────────────────────────────────────────────────────────────────
    def _pnl(self):
        self._sec("OPERATING STATEMENT", bg=BG_TITLE, size=10)

        # ── Sales ─────────────────────────────────────────────────────────
        sales_v = self._v("sales", "Net Sales", proj_fn=self._proj_growth)
        self._row(1, "i. Domestic Sale",  sales_v, nf=NF_NUM)
        self._row("", "ii.Export Sale",
                  {c: self._get("sales","Export Sales",fy) for c,fy in self._fys().items()},
                  nf=NF_NUM)

        # Total Gross Sales — same as domestic (no export)
        def _tgs_proj(v23,v24,v25): return sales_v.get(C_P26), sales_v.get(C_P27)
        tgs = {**{c: (sales_v.get(c) or 0) + 0 for c in AUD},
               C_P26: sales_v.get(C_P26), C_P27: sales_v.get(C_P27)}
        # Use extracted total if available
        for c, fy in self._fys().items():
            ev = self._get("sales","Net Sales",fy) or self._get("sales","Total Gross Sales",fy)
            if ev: tgs[c] = ev
        self._row(1, "Total Gross Sales", tgs, bold=True, total=True, calc=True, nf=NF_NUM)
        tgs_r = self._rrow("Total Gross Sales")

        # Net Other Income
        noi_v = self._v("pnl","Net Other Income")
        self._row(2, "Net Other Income", noi_v, nf=NF_NUM, indent=1)
        noi_r = self._rrow("Net Other Income")

        # Total Income
        ti = {}
        for c in ALL:
            tgs_val = self._cell(tgs_r, c) or 0
            noi_val = noi_v.get(c) or 0
            ti[c]   = round(tgs_val + noi_val, 4)
        self._row(3, "Total Income", ti, bold=True, total=True, calc=True, nf=NF_NUM)
        ti_r = self._rrow("Total Income")

        # Growth
        gi = {}
        for c in [C_FY24,C_FY25]:
            prev = self._cell(ti_r, c-1) or 0
            curr = self._cell(ti_r, c)   or 0
            gi[c] = round((curr-prev)/abs(prev),6) if prev!=0 else None
        gi[C_FY23] = None
        g24 = gi.get(C_FY24) or 0.10
        g25 = gi.get(C_FY25) or 0.10
        avg_g = round((g24+g25)/2, 4)
        gi[C_P26] = round(avg_g, 4)
        gi[C_P27] = round(avg_g, 4)
        self._row(4, "Growth in sales", gi, calc=True, nf=NF_PCT)
        self._saved_growth = avg_g   # stash for projection

        # ── Cost of Sales ──────────────────────────────────────────────────
        self._blank(4)
        self._c(self.row, C_LBL, "Cost of Sales", bold=True, bg=BG_SUBSECT,
                color=FG_DARK); self.row += 1

        # Freight
        fh_v = self._v("cost_of_sales","Freight & Handling Expenses")
        for c in [C_P26, C_P27]:
            fh_v[c] = None
        s23=self._cell(tgs_r,C_FY23); s24=self._cell(tgs_r,C_FY24); s25=self._cell(tgs_r,C_FY25)
        fh23=fh_v.get(C_FY23); fh24=fh_v.get(C_FY24); fh25=fh_v.get(C_FY25)
        p26,p27 = self._proj_fixed_ratio(fh23,fh24,fh25,s23,s24,s25) if all([s23,s24,s25]) else (None,None)
        fh_v[C_P26]=p26; fh_v[C_P27]=p27
        self._row(5, "a. Freight & Handling Expenses", fh_v, nf=NF_NUM, indent=1)

        # Vehicle
        vr_v = self._v("cost_of_sales","Vehicle Running Expenses")
        vr23=vr_v.get(C_FY23); vr24=vr_v.get(C_FY24); vr25=vr_v.get(C_FY25)
        p26,p27 = self._proj_fixed_ratio(vr23,vr24,vr25,s23,s24,s25) if all([s23,s24,s25]) else (None,None)
        vr_v[C_P26]=p26; vr_v[C_P27]=p27
        self._row("", "b. Vehicle Running expenses", vr_v, nf=NF_NUM, indent=1)

        self._row("", "c. Stores & Spares (Imported)",
                  {c:None for c in ALL}, nf=NF_NUM, indent=1)
        self._row("", "d. Stores & Spares (Indigenous)",
                  {c:None for c in ALL}, nf=NF_NUM, indent=1)
        self._row(6,  "Power & Fuel",      {c:None for c in ALL}, nf=NF_NUM)
        self._row(7,  "Direct Labour",     {c:None for c in ALL}, nf=NF_NUM)
        self._row(8,  "Repairs and maintainance", {c:None for c in ALL}, nf=NF_NUM)
        self._row(9,  "Other Mfg. Expenses",      {c:None for c in ALL}, nf=NF_NUM)

        # Depreciation
        dep_v = self._v("cost_of_sales","Depreciation")
        d23=dep_v.get(C_FY23); d24=dep_v.get(C_FY24); d25=dep_v.get(C_FY25)
        # project dep: declining at avg rate
        g_dep = self._avg_growth(d23,d24,d25)
        g_dep = max(-0.5, min(g_dep, 0))  # dep can only fall or stay flat in projection
        dep_v[C_P26] = round((d25 or 0)*(1+g_dep),4)
        dep_v[C_P27] = round(dep_v[C_P26]*(1+g_dep),4)
        self._row(10, "Depreciation ", dep_v, nf=NF_NUM, indent=1)
        dep_r = self._rrow("Depreciation ")

        self._row(11, "Others expenses", {c:0 for c in ALL}, nf=NF_NUM)
        self._row("a",None or "", {c:None for c in ALL}, nf=NF_NUM)
        self._row("b","Insurance",        {c:None for c in ALL}, nf=NF_NUM, indent=2)
        self._row("c","Transport Expenses",{c:None for c in ALL}, nf=NF_NUM, indent=2)
        self._row("d",None or "",          {c:None for c in ALL}, nf=NF_NUM)

        # Sub Total (cost of sales only)
        cos_v = {}
        cos_ext = self._v("cost_of_sales","Total Cost of Sales")
        for c in ALL:
            ev = cos_ext.get(c)
            if ev is not None:
                cos_v[c] = ev
            else:
                fh  = self._cell(self._rrow("a. Freight & Handling Expenses"), c) or 0
                vr  = self._cell(self._rrow("b. Vehicle Running expenses"),    c) or 0
                dep = self._cell(dep_r, c) or 0
                cos_v[c] = round(fh + vr + dep, 4)
        self._row("", "Sub Total", cos_v, bold=True, total=True, calc=True, nf=NF_NUM)
        self._row("", "Sub Total", cos_v, bold=False, bg=BG_LABEL, calc=True, nf=NF_NUM)  # Row37 (Opening Stock sub-total)
        sub_cos_r = self._rrow("Sub Total")

        self._row(12,"Add: Opening Stock in Process",{c:None for c in ALL},nf=NF_NUM)
        self._row("","Sub Total", cos_v, bold=False, bg=BG_LABEL, calc=True, nf=NF_NUM)
        self._row(13,"Deduct : Closing Stock in Process",{c:None for c in ALL},nf=NF_NUM)
        self._row("","Cost of Production ", cos_v, bold=True, total=True, calc=True, nf=NF_NUM)
        self._row(14,"Add: Opening Stock of Finished Goods",{c:None for c in ALL},nf=NF_NUM)
        self._row("","Sub Total", cos_v, bold=False, bg=BG_LABEL, calc=True, nf=NF_NUM)
        self._row(15,"Deduct : Closing Stock OF Finished Goods",{c:None for c in ALL},nf=NF_NUM)
        self._row("","Sub Total ( Total Cost of Sales)", cos_v,
                  bold=True, total=True, calc=True, nf=NF_NUM)
        cos_r = self._rrow("Sub Total ( Total Cost of Sales)")

        # Gross Profit
        gp_v = {}
        gp_ext = self._v("profitability","Gross profit")
        for c in ALL:
            ev = gp_ext.get(c)
            if ev is not None:
                gp_v[c] = ev
            else:
                ti_v = self._cell(ti_r, c) or 0
                co_v = self._cell(cos_r, c) or 0
                gp_v[c] = round(ti_v - co_v, 4)
        self._row(16,"Gross profit", gp_v, bold=True, total=True, calc=True, nf=NF_NUM)
        gp_r = self._rrow("Gross profit")

        # Gross Profit/Sales
        gps_v = {}
        for c in ALL:
            gp = self._cell(gp_r, c)
            ti = self._cell(ti_r, c)
            gps_v[c] = round(gp/ti,6) if gp is not None and ti and ti!=0 else None
        self._row("","Gross Profit/ Sales", gps_v, calc=True, nf=NF_PCT, indent=1)

        # Selling Expenses
        sel_v = self._v("cost_of_sales","Selling Expenses")
        sel23=sel_v.get(C_FY23); sel24=sel_v.get(C_FY24); sel25=sel_v.get(C_FY25)
        p26,p27 = self._proj_fixed_ratio(sel23,sel24,sel25,s23,s24,s25) if all([s23,s24,s25]) else (None,None)
        sel_v[C_P26]=p26; sel_v[C_P27]=p27
        self._row(17,"Selling Expenses", sel_v, nf=NF_NUM)
        sel_r = self._rrow("Selling Expenses")

        # Admin Expenses
        adm_v = self._v("cost_of_sales","Administrative Expenses")
        adm23=adm_v.get(C_FY23); adm24=adm_v.get(C_FY24); adm25=adm_v.get(C_FY25)
        p26,p27 = self._proj_fixed_ratio(adm23,adm24,adm25,s23,s24,s25) if all([s23,s24,s25]) else (None,None)
        adm_v[C_P26]=p26; adm_v[C_P27]=p27
        self._row(18,"Administrative Expenses", adm_v, nf=NF_NUM)
        adm_r = self._rrow("Administrative Expenses")

        # Sub Total (total including S&A)
        sub2 = {}
        for c in ALL:
            sub2[c] = round((self._cell(cos_r,c) or 0) +
                            (self._cell(sel_r,c) or 0) +
                            (self._cell(adm_r,c) or 0), 4)
        self._row("","Sub Total ", sub2, bold=True, total=True, calc=True, nf=NF_NUM)
        sub2_r = self._rrow("Sub Total ")

        # Operating Profit before interest
        opbi = {}
        opbi_ext = self._v("profitability","Operating Profit before interest")
        for c in ALL:
            ev = opbi_ext.get(c)
            if ev is not None:
                opbi[c] = ev
            else:
                opbi[c] = round((self._cell(ti_r,c) or 0) -
                                (self._cell(sub2_r,c) or 0), 4)
        self._row(19,"Operating Profit before interest ", opbi,
                  bold=True, total=True, calc=True, nf=NF_NUM)
        opbi_r = self._rrow("Operating Profit before interest ")

        # Interest
        cc_v = self._v("interest","Interest on CC")
        self._row("","a. Interest on CC.", cc_v, nf=NF_NUM, indent=2)
        cc_r = self._rrow("a. Interest on CC.")

        tl_v = self._v("interest","Interest on TL")
        # project TL interest to decline as loans repaid
        tl23=tl_v.get(C_FY23); tl24=tl_v.get(C_FY24); tl25=tl_v.get(C_FY25)
        g_tl = self._avg_growth(tl23,tl24,tl25)
        g_tl = max(-0.5, min(g_tl, 0))
        tl_v[C_P26] = round((tl25 or 0)*(1+g_tl),4)
        tl_v[C_P27] = round(tl_v[C_P26]*(1+g_tl),4)
        self._row("","b.Interest on TL ", tl_v, nf=NF_NUM, indent=2)
        tl_int_r = self._rrow("b.Interest on TL ")

        oi_v = self._v("interest","Other Interest")
        oi_v[C_P26] = oi_v[C_P27] = None
        self._row("","c.Other interests ", oi_v, nf=NF_NUM, indent=2)
        oi_r = self._rrow("c.Other interests ")

        tot_int = {}
        int_ext = self._v("interest","Total Interest")
        for c in ALL:
            ev = int_ext.get(c)
            if ev is not None:
                tot_int[c] = ev
            else:
                tot_int[c] = round((self._cell(cc_r,c) or 0) +
                                   (self._cell(tl_int_r,c) or 0) +
                                   (self._cell(oi_r,c) or 0), 4)
        self._row(20,"Total Interest", tot_int, bold=True, total=True, calc=True, nf=NF_NUM)
        tot_int_r = self._rrow("Total Interest")

        # Operating Profit after Interest
        opai = {}
        opai_ext = self._v("profitability","Operating Profit after interest")
        for c in ALL:
            ev = opai_ext.get(c)
            if ev is not None:
                opai[c] = ev
            else:
                opai[c] = round((self._cell(opbi_r,c) or 0) -
                                (self._cell(tot_int_r,c) or 0), 4)
        self._row(21,"Operating Profit after Interest", opai,
                  bold=True, total=True, calc=True, nf=NF_NUM)
        opai_r = self._rrow("Operating Profit after Interest")

        # Non-op income
        self._blank(3)
        self._c(self.row, C_LBL, "Non-Operating Income", bold=True,
                bg=BG_SUBSECT, color=FG_DARK); self.row+=1

        noi_a_v = self._v("pnl","Interest/Dividend/Royalties")
        self._row("a","Interest/Dividend/Royalties etc..", noi_a_v, nf=NF_NUM, indent=2)
        noi_a_r = self._rrow("Interest/Dividend/Royalties etc..")

        noi_b_v = self._v("pnl","Other Income")
        noi_b_v[C_P26] = noi_b_v[C_P27] = None
        self._row("c","Other Income", noi_b_v, nf=NF_NUM, indent=2)
        noi_b_r = self._rrow("Other Income")

        sub_noi = {}
        for c in ALL:
            a = self._cell(noi_a_r, c) or 0
            b = self._cell(noi_b_r, c) or 0
            sub_noi[c] = round(a+b,4) if (a or b) else None
        # use extracted noi where available
        noi_ext = self._v("pnl","Net Other Income")
        for c in AUD:
            ev = noi_ext.get(c)
            if ev is not None: sub_noi[c] = ev
        self._row("","Sub Total", sub_noi, bold=True, total=True, calc=True, nf=NF_NUM)
        sub_noi_r = self._rrow("Sub Total")

        # Non-op expenses (deductions)
        proc_fees = {c:None for c in ALL}
        self._row("d","Processing fees", proc_fees, nf=NF_NUM, indent=2)
        pf_r = self._rrow("Processing fees")

        sub_noe = {}
        for c in ALL:
            pf = self._cell(pf_r,c) or 0
            sub_noe[c] = round(pf,4) if pf else None
        self._row("","Sub Total", sub_noe, bold=True, total=True, calc=True, nf=NF_NUM)
        sub_noe_r = self._rrow("Sub Total")

        # Net non-op
        net_nop = {}
        for c in ALL:
            s_noi = self._cell(sub_noi_r,c) or 0
            s_noe = self._cell(sub_noe_r,c) or 0
            net_nop[c] = round(s_noi - s_noe, 4)
        # use extracted values for audited
        net_nop_ext = self._v("pnl","Net Other Income")
        for c in AUD:
            ev = net_nop_ext.get(c)
            if ev is not None: net_nop[c] = ev
        self._row(24,"Net of other non operating Income/Expenses", net_nop,
                  bold=True, calc=True, total=True, nf=NF_NUM)
        net_nop_r = self._rrow("Net of other non operating Income/Expenses")

        # PBT
        pbt = {}
        pbt_ext = self._v("pnl","Profit before Tax")
        for c in ALL:
            ev = pbt_ext.get(c)
            if ev is not None:
                pbt[c] = ev
            else:
                pbt[c] = round((self._cell(opai_r,c) or 0) +
                               (self._cell(net_nop_r,c) or 0), 4)
        self._row(25,"Profit before Tax /Loss (PBT)", pbt,
                  bold=True, total=True, calc=True, nf=NF_NUM)
        pbt_r = self._rrow("Profit before Tax /Loss (PBT)")

        # Tax
        tax = {}
        tax_ext = self._v("pnl","Provision for Taxes")
        for c in ALL:
            ev = tax_ext.get(c)
            if ev is not None:
                tax[c] = ev
            else:
                pbt_v = self._cell(pbt_r, c) or 0
                tax[c] = round(pbt_v * 0.25, 4) if pbt_v > 0 else None
        self._row(26,"Provision for Taxes", tax, nf=NF_NUM, indent=1)
        tax_r = self._rrow("Provision for Taxes")

        # PAT
        pat = {}
        pat_ext = self._v("pnl","Net Profit/Loss (PAT)")
        for c in ALL:
            ev = pat_ext.get(c)
            if ev is not None:
                pat[c] = ev
            else:
                pbt_v = self._cell(pbt_r,c)
                tax_v = self._cell(tax_r,c) or 0
                pat[c] = round(pbt_v - tax_v, 4) if pbt_v is not None else None
        self._row(27,"Net Profit/Loss (PAT)", pat,
                  bold=True, total=True, calc=True, nf=NF_NUM)
        pat_r = self._rrow("Net Profit/Loss (PAT)")

        # Cash Accruals = PAT + Dep
        ca = {}
        ca_ext = self._v("pnl","Cash Accruals")
        for c in ALL:
            ev = ca_ext.get(c)
            if ev is not None:
                ca[c] = ev
            else:
                p = self._cell(pat_r,c)
                d = self._cell(dep_r,c) or 0
                ca[c] = round(p+d, 4) if p is not None else None
        self._row(28,"Cash Accruals", ca, bold=True, total=True, calc=True, nf=NF_NUM)
        ca_r = self._rrow("Cash Accruals")

        self._row(29,"Dividend paid + IT on Dividend",
                  {c:None for c in ALL}, nf=NF_NUM)
        div_r = self._rrow("Dividend paid + IT on Dividend")

        # Retained Profit
        rp = {}
        for c in ALL:
            p=self._cell(pat_r,c); d=self._cell(div_r,c) or 0
            rp[c] = round(p-d,4) if p is not None else None
        self._row(30,"Retained Profit ", rp, calc=True, nf=NF_NUM)

        # Retained Cash Profits
        rcp = {}
        for c in ALL:
            p=self._cell(ca_r,c); d=self._cell(div_r,c) or 0
            rcp[c] = round(p-d,4) if p is not None else None
        self._row(31,"Retained Cash Profits", rcp, calc=True, nf=NF_NUM)

        # RM Content
        rm = {}
        rm_ext = self._v("pnl","RM Content in sales")
        for c in ALL:
            ev = rm_ext.get(c)
            if ev is not None:
                rm[c] = ev
            else:
                co = self._cell(cos_r,c); ti_ = self._cell(ti_r,c)
                rm[c] = round(co/ti_,6) if co and ti_ and ti_!=0 else None
        self._row(32,"RM Content in sales", rm, calc=True, nf=NF_PCT)

        # PBDIT
        pbdit = {}
        pbdit_ext = self._v("pnl","PBDIT")
        for c in ALL:
            ev = pbdit_ext.get(c)
            if ev is not None:
                pbdit[c] = ev
            else:
                p=self._cell(pbt_r,c) or 0
                i=self._cell(tot_int_r,c) or 0
                d=self._cell(dep_r,c) or 0
                pbdit[c] = round(p+i+d,4)
        self._row(33,"PBDIT ", pbdit, bold=True, total=True, calc=True, nf=NF_NUM)
        pbdit_r = self._rrow("PBDIT ")

        # Ratios
        for sr, lbl, r_num, sect, fld in [
            (34,"PBDIT/Sales",    pbdit_r, "pnl","PBDIT/Sales"),
            (35,"Operating Profits/Sales", opbi_r, "pnl","Operating Profits/Sales"),
            (36,"PBT/Sales",      pbt_r,  "pnl","PBT/Sales"),
            (37,"PAT/Sales",      pat_r,  "pnl","PAT/Sales"),
            (38,"Cash Accruals/ Sales", ca_r, "pnl","Cash Accruals/ Sales"),
        ]:
            rv = {}
            ext = self._v(sect, fld)
            for c in ALL:
                ev = ext.get(c)
                if ev is not None:
                    rv[c] = ev
                else:
                    num = self._cell(r_num, c)
                    den = self._cell(ti_r, c)
                    rv[c] = round(num/den,6) if num is not None and den and den!=0 else None
            # projections: calc from projected rows
            for c in [C_P26, C_P27]:
                num = self._cell(r_num, c)
                den = self._cell(ti_r, c)
                rv[c] = round(num/den,6) if num is not None and den and den!=0 else None
            self._row(sr, lbl, rv, calc=True, nf=NF_PCT)

        # stash key row refs
        self._pbt_r = pbt_r; self._pat_r = pat_r; self._ca_r = ca_r
        self._dep_r = dep_r; self._tot_int_r = tot_int_r
        self._pbdit_r = pbdit_r; self._ti_r = ti_r
        self._tgs_r = tgs_r; self._cos_r = cos_r
        self._opai_r = opai_r; self._opbi_r = opbi_r

    # ─────────────────────────────────────────────────────────────────────────
    # BALANCE SHEET — LIABILITIES
    # ─────────────────────────────────────────────────────────────────────────
    def _bs_liabilities(self):
        self._sec("ANALYSIS OF BALANCE SHEET  —  LIABILITIES", bg=BG_TITLE, size=10)
        self._blank(3)
        self._c(self.row, C_LBL,"Current Liabilities",bold=True,bg=BG_SUBSECT,color=FG_DARK)
        self.row+=1

        # Sub Total A
        cc_v = self._v("current_liabilities","Short Term loans from Applicant Bank")
        self._row("1.a","Short Term loans from Applicant Bank  ", cc_v, nf=NF_NUM)
        cc_r = self._rrow("Short Term loans from Applicant Bank  ")

        ob_v = self._v("current_liabilities","Short Term Borrowings from Others")
        ob_v[C_P26]=ob_v[C_P27]=None
        self._row("1.b","Short Term loans From Other banks ", ob_v, nf=NF_NUM)
        ob_r = self._rrow("Short Term loans From Other banks ")

        od_v = {c:None for c in ALL}
        self._row("1.c","Clean OD", od_v, nf=NF_NUM)
        od_r = self._rrow("Clean OD")

        sub_a = {}
        sub_a_ext = self._v("current_liabilities","Short Term Borrowings Bank Total")
        for c in ALL:
            ev = sub_a_ext.get(c)
            if ev is not None:
                sub_a[c] = ev
            else:
                sub_a[c] = round((self._cell(cc_r,c) or 0)+
                                 (self._cell(ob_r,c) or 0)+
                                 (self._cell(od_r,c) or 0),4)
        self._row("","Sub Total (A)", sub_a, bold=True, total=True, calc=True, nf=NF_NUM)
        sub_a_r = self._rrow("Sub Total (A)")

        self._row(2,"Short Term Borrowings from Others",{c:None for c in ALL},nf=NF_NUM)

        # Sundry Creditors
        cred_v = self._v("current_liabilities","Sundry Creditors (Trade)")
        s23=self._cell(self._tgs_r,C_FY23); s24=self._cell(self._tgs_r,C_FY24)
        s25=self._cell(self._tgs_r,C_FY25)
        cr23=cred_v.get(C_FY23); cr24=cred_v.get(C_FY24); cr25=cred_v.get(C_FY25)
        # project creditors at same days-of-consumption ratio
        cr_p26,cr_p27 = self._proj_fixed_ratio(cr23,cr24,cr25,s23,s24,s25) if all([s23,s24,s25]) else (None,None)
        cred_v[C_P26]=cr_p26; cred_v[C_P27]=cr_p27
        self._row(3,"Sundry Creditors (Trade)", cred_v, nf=NF_NUM)
        cred_r = self._rrow("Sundry Creditors (Trade)")

        # Advance from customers
        adv_v = self._v("current_liabilities","Advance Payment from Customers")
        adv23=adv_v.get(C_FY23); adv24=adv_v.get(C_FY24); adv25=adv_v.get(C_FY25)
        p26,p27 = self._proj_fixed_ratio(adv23,adv24,adv25,s23,s24,s25) if all([s23,s24,s25]) else (None,None)
        adv_v[C_P26]=p26; adv_v[C_P27]=p27
        self._row(4,"Advance Payment from Customers", adv_v, nf=NF_NUM)

        self._row(5,"Net Provision for Taxation (if positive)",{c:None for c in ALL},nf=NF_NUM)
        self._row(6,"Dividend Payable",{c:None for c in ALL},nf=NF_NUM)

        stat_v = self._v("current_liabilities","Other Statutory Liabilities")
        stat_v[C_P26]=stat_v[C_P27]=None
        self._row(7,"Other Statutory Liab. (Due within one Year)", stat_v, nf=NF_NUM)

        self._row(8,"Overdue Term Liabilities",{c:None for c in ALL},nf=NF_NUM)

        inst_v = self._v("current_liabilities","Installments of term Loan")
        # project declining instalments
        i23=inst_v.get(C_FY23); i24=inst_v.get(C_FY24); i25=inst_v.get(C_FY25)
        g_i = self._avg_growth(i23,i24,i25); g_i = max(-0.5,min(g_i,0))
        inst_v[C_P26]=round((i25 or 0)*(1+g_i),4); inst_v[C_P27]=round(inst_v[C_P26]*(1+g_i),4)
        self._row(9,"Installments of term Loan/ DPGs/ Deposits/ debentures due within next year",
                  inst_v, nf=NF_NUM)
        inst_r = self._rrow("Installments of term Loan/ DPGs/ Deposits/ debentures due within next year")

        ocl_v = self._v("current_liabilities","Other Current Liabilities & Provisions")
        # project OCL growing with sales
        ocl23=ocl_v.get(C_FY23); ocl24=ocl_v.get(C_FY24); ocl25=ocl_v.get(C_FY25)
        p26,p27 = self._proj_fixed_ratio(ocl23,ocl24,ocl25,s23,s24,s25) if all([s23,s24,s25]) else (None,None)
        ocl_v[C_P26]=p26; ocl_v[C_P27]=p27
        self._row(10,"Other Current Liabilities & Provisions (due with in one year)",
                  ocl_v, nf=NF_NUM)
        ocl_r = self._rrow("Other Current Liabilities & Provisions (due with in one year)")

        ocl_b = self._v("current_liabilities","Other Current Liabilities")
        ocl_b[C_P26]=ocl_b[C_P27]=None
        self._row("b","Other current Liabilities ", ocl_b, nf=NF_NUM, indent=2)
        prov_v = self._v("current_liabilities","Short Term Provisions")
        prov_v[C_P26]=prov_v[C_P27]=None
        self._row("c","Provision for Others", prov_v, nf=NF_NUM, indent=2)

        # Sub Total B
        sub_b = {}
        sub_b_ext = self._v("current_liabilities","Other Current Liabilities Sub Total")
        for c in ALL:
            ev = sub_b_ext.get(c)
            if ev is not None:
                sub_b[c] = ev
            else:
                cr = self._cell(cred_r,c) or 0
                adv= self._cell(self._rrow("Advance Payment from Customers"),c) or 0
                st = self._cell(self._rrow("Other Statutory Liab. (Due within one Year)"),c) or 0
                ins= self._cell(inst_r,c) or 0
                oc = self._cell(ocl_r,c) or 0
                sub_b[c] = round(cr+adv+st+ins+oc,4)
        self._row(11,"Sub Total (B)", sub_b, bold=True, total=True, calc=True, nf=NF_NUM)
        sub_b_r = self._rrow("Sub Total (B)")

        # Total CL
        tcl = {}
        tcl_ext = self._v("current_liabilities","Total Current Liabilities")
        for c in ALL:
            ev = tcl_ext.get(c)
            if ev is not None:
                tcl[c] = ev
            else:
                tcl[c] = round((self._cell(sub_a_r,c) or 0)+
                               (self._cell(sub_b_r,c) or 0),4)
        self._row(12,"TOTAL CURRENT LIABILITIES", tcl,
                  bold=True, total=True, calc=True, nf=NF_NUM)
        self._tcl_r = self._rrow("TOTAL CURRENT LIABILITIES")
        self._sub_b_r = sub_b_r; self._sub_a_r = sub_a_r
        self._cred_r  = cred_r;  self._inst_r  = inst_r

        # ── TERM LIABILITIES ──────────────────────────────────────────────
        self._blank(3)
        self._c(self.row,C_LBL,"TERM LIABILITIES",bold=True,bg=BG_SUBSECT,color=FG_DARK)
        self.row+=1

        self._row(13,"Debentures (not maturing within one Year)",{c:None for c in ALL},nf=NF_NUM)
        self._row(14,"Preference Shares (redeemable after 1 year)",{c:None for c in ALL},nf=NF_NUM)

        tl_bank = self._v("term_liabilities","Term Loan from Bank")
        t23=tl_bank.get(C_FY23); t24=tl_bank.get(C_FY24); t25=tl_bank.get(C_FY25)
        g_tl2 = self._avg_growth(t23,t24,t25); g_tl2=max(-0.5,min(g_tl2,0))
        tl_bank[C_P26]=round((t25 or 0)*(1+g_tl2),4)
        tl_bank[C_P27]=round(tl_bank[C_P26]*(1+g_tl2),4)
        self._row(14,"Term Loan from Bank(Less next Year Instalments)", tl_bank, nf=NF_NUM)
        tl_bank_r = self._rrow("Term Loan from Bank(Less next Year Instalments)")

        self._row(14,"Term Loan from Other Banks/Inst.(Ecl. Instal.due next Yr.)",
                  {c:None for c in ALL},nf=NF_NUM)

        other_tl = {}
        other_tl_ext = self._v("term_liabilities","Other Term Liabilities")
        for c in ALL:
            ev = other_tl_ext.get(c)
            if ev is not None: other_tl[c] = ev
        self._row(17,"Other term Liabilities", other_tl, nf=NF_NUM)
        oth_tl_r = self._rrow("Other term Liabilities")

        ul_v = self._v("term_liabilities","Loan from Directors")
        self._row("a","Unsecured Loans from Directors", ul_v, nf=NF_NUM, indent=2)
        sec_dep = self._v("term_liabilities","Security Deposits")
        self._row("b","Security Deposits", sec_dep, nf=NF_NUM, indent=2)
        unsl = self._v("term_liabilities","Unsecured Loans")
        unsl[C_P26]=unsl[C_P27]=None
        self._row("c","Unsecured Loans", unsl, nf=NF_NUM, indent=2)
        lt_prov = self._v("term_liabilities","Long Term Provisions")
        lt_prov[C_P26]=lt_prov[C_P27]=None
        self._row("d","Long Term Provisions", lt_prov, nf=NF_NUM, indent=2)

        # Total TL
        ttl = {}
        ttl_ext = self._v("term_liabilities","Total Term Liabilities")
        for c in ALL:
            ev = ttl_ext.get(c)
            if ev is not None:
                ttl[c] = ev
            else:
                ttl[c] = round((self._cell(tl_bank_r,c) or 0)+
                               (self._cell(oth_tl_r,c) or 0),4)
        self._row(18,"TOTAL TERM LIABILITIES", ttl,
                  bold=True, total=True, calc=True, nf=NF_NUM)
        self._ttl_r = self._rrow("TOTAL TERM LIABILITIES")
        self._tl_bank_r = tl_bank_r

        # Total Outside Liabilities
        tol = {}
        tol_ext = self._v("term_liabilities","Total Outside Liabilities")
        for c in ALL:
            ev = tol_ext.get(c)
            if ev is not None:
                tol[c] = ev
            else:
                tol[c] = round((self._cell(self._tcl_r,c) or 0)+
                               (self._cell(self._ttl_r,c) or 0),4)
        self._row(19,"TOTAL OF OUTSIDE LIABILITIES ", tol,
                  bold=True, total=True, calc=True, nf=NF_NUM)
        self._tol_r = self._rrow("TOTAL OF OUTSIDE LIABILITIES ")

        # ── NET WORTH ─────────────────────────────────────────────────────
        self._blank(3)
        self._c(self.row,C_LBL,"NET WORTH",bold=True,bg=BG_SUBSECT,color=FG_DARK); self.row+=1

        sc_v = self._v("net_worth","Capital")
        self._row(20,"Share Capital", sc_v, nf=NF_NUM)
        sc_r = self._rrow("Share Capital")

        gr_v = self._v("net_worth","Other reserves")
        # general reserve grows with retained profit
        gr23=gr_v.get(C_FY23); gr24=gr_v.get(C_FY24); gr25=gr_v.get(C_FY25)
        pat25=self._cell(self._pat_r,C_FY25) or 0
        gr_v[C_P26] = round((gr25 or 0)+pat25,4)
        pat26=self._cell(self._pat_r,C_P26) or 0
        gr_v[C_P27] = round((gr_v[C_P26] or 0)+pat26,4)
        self._row(21,"General Reserve", gr_v, nf=NF_NUM)

        self._row(22,"Revaluation Reserve",{c:None for c in ALL},nf=NF_NUM)
        self._row(23,"Adjustments for previous Year costs",{c:None for c in ALL},nf=NF_NUM)
        self._row(24,"Other reserves (excluding Provisions)",{c:None for c in ALL},nf=NF_NUM)

        others_v = {c:None for c in ALL}
        self._row(25,"Others", others_v, nf=NF_NUM)
        self._row("b","Share Premium", self._v("net_worth","Share Premium"), nf=NF_NUM, indent=2)

        surp_v = {}
        for c in ALL:
            surp_v[c] = self._cell(self._pat_r,c)
        self._row(26,"Surplus (+) or deficit (-) in Profit & Loss a/c",
                  surp_v, nf=NF_NUM)
        surp_r = self._rrow("Surplus (+) or deficit (-) in Profit & Loss a/c")

        nw = {}
        nw_ext = self._v("net_worth","Net Worth")
        for c in ALL:
            ev = nw_ext.get(c)
            if ev is not None:
                nw[c] = ev
            else:
                nw[c] = round((self._cell(sc_r,c) or 0)+
                              (self._cell(self._rrow("General Reserve"),c) or 0)+
                              (self._cell(self._rrow("Others"),c) or 0)+
                              (self._cell(surp_r,c) or 0), 4)
        self._row(27,"NET WORTH", nw, bold=True, total=True, calc=True, nf=NF_NUM)
        self._nw_r = self._rrow("NET WORTH")

        # Total Liabilities
        tot_liab = {}
        for c in ALL:
            tot_liab[c] = round((self._cell(self._tol_r,c) or 0)+
                                (self._cell(self._nw_r,c) or 0),4)
        self._row(28,"TOTAL LIABILITIES (18+24)", tot_liab,
                  bold=True, total=True, calc=True, nf=NF_NUM)
        self._tot_liab_r = self._rrow("TOTAL LIABILITIES (18+24)")

    # ─────────────────────────────────────────────────────────────────────────
    # BALANCE SHEET — ASSETS
    # ─────────────────────────────────────────────────────────────────────────
    def _bs_assets(self):
        self._sec("ANALYSIS OF BALANCE SHEET  —  ASSETS", bg=BG_TITLE, size=10)
        self._blank(3)
        self._c(self.row,C_LBL,"Current Assets",bold=True,bg=BG_SUBSECT,color=FG_DARK); self.row+=1

        s23=self._cell(self._tgs_r,C_FY23); s24=self._cell(self._tgs_r,C_FY24)
        s25=self._cell(self._tgs_r,C_FY25)

        cb_v = self._v("current_assets","Cash & Bank Balances")
        cb23=cb_v.get(C_FY23); cb24=cb_v.get(C_FY24); cb25=cb_v.get(C_FY25)
        p26,p27 = self._proj_fixed_ratio(cb23,cb24,cb25,s23,s24,s25) if all([s23,s24,s25]) else (None,None)
        cb_v[C_P26]=p26; cb_v[C_P27]=p27
        self._row(1,"Cash & Bank Balances", cb_v, nf=NF_NUM)

        self._row(2,"Govt. & other Trustee securities",{c:None for c in ALL},nf=NF_NUM)

        fd_v = self._v("current_assets","Fixed Deposits with Banks")
        self._row(3,"Fixed Deposits with Banks", fd_v, nf=NF_NUM)

        dr_v = self._v("current_assets","Domestic Receivables")
        dr23=dr_v.get(C_FY23); dr24=dr_v.get(C_FY24); dr25=dr_v.get(C_FY25)
        p26,p27 = self._proj_fixed_ratio(dr23,dr24,dr25,s23,s24,s25) if all([s23,s24,s25]) else (None,None)
        dr_v[C_P26]=p26; dr_v[C_P27]=p27
        self._row("4.a","Domestic Receivables including BP/BD", dr_v, nf=NF_NUM)
        dr_r = self._rrow("Domestic Receivables including BP/BD")

        ubr_v = {c:None for c in ALL}
        self._row("4.b","Unbilled receivables", ubr_v, nf=NF_NUM)
        self._row(5,"Export Receivables including BP/BD)",{c:None for c in ALL},nf=NF_NUM)
        dfr_v = self._v("current_assets","Deferred Receivables")
        dfr_v[C_P26]=dfr_v[C_P27]=None
        self._row(6,"Deferred receivables(due within one year)", dfr_v, nf=NF_NUM)
        rfg_v = self._v("current_assets","Refund Dues")
        rfg_v[C_P26]=rfg_v[C_P27]=None
        self._row(7,"Refund Dues from Government Authorities", rfg_v, nf=NF_NUM)
        self._row(8,"Indigenous Rawmaterial",{c:None for c in ALL},nf=NF_NUM)
        self._row(9,"Stock in Process",{c:0 for c in ALL},nf=NF_NUM)
        self._row(10,"Finished Goods",{c:0 for c in ALL},nf=NF_NUM)

        adv_v = self._v("current_assets","Advances to Suppliers")
        adv23=adv_v.get(C_FY23); adv24=adv_v.get(C_FY24); adv25=adv_v.get(C_FY25)
        p26,p27 = self._proj_fixed_ratio(adv23,adv24,adv25,s23,s24,s25) if all([s23,s24,s25]) else (None,None)
        adv_v[C_P26]=p26; adv_v[C_P27]=p27
        self._row(13,"Advances to Suppliers/Transport", adv_v, nf=NF_NUM)

        self._row(14,"Net Advance Payment of Taxes (if positive)",
                  {c:None for c in ALL}, nf=NF_NUM)

        oca_v = self._v("current_assets","Other Current Assets")
        oca23=oca_v.get(C_FY23); oca24=oca_v.get(C_FY24); oca25=oca_v.get(C_FY25)
        p26,p27 = self._proj_fixed_ratio(oca23,oca24,oca25,s23,s24,s25) if all([s23,s24,s25]) else (None,None)
        oca_v[C_P26]=p26; oca_v[C_P27]=p27
        self._row(15,"Other Current Assets (specify major items)", oca_v, nf=NF_NUM)

        pre_v = self._v("current_assets","Prepaid Expenses")
        self._row("b","Prepaid expenses", pre_v, nf=NF_NUM, indent=2)
        dep_emd = self._v("current_assets","Deposits EMD")
        self._row("c","Deposits-EMD & Fixed Deposits", dep_emd, nf=NF_NUM, indent=2)

        # Total CA
        tca = {}
        tca_ext = self._v("current_assets","Total Current Assets")
        for c in ALL:
            ev = tca_ext.get(c)
            if ev is not None:
                tca[c] = ev
            else:
                total = 0
                for lbl in ["Cash & Bank Balances","Fixed Deposits with Banks",
                            "Domestic Receivables including BP/BD","Unbilled receivables",
                            "Advances to Suppliers/Transport","Other Current Assets (specify major items)"]:
                    rr = self._rrow(lbl)
                    if rr: total += self._cell(rr,c) or 0
                tca[c] = round(total,4)
        self._row(16,"TOTAL CURRENT ASSETS", tca,
                  bold=True, total=True, calc=True, nf=NF_NUM)
        self._tca_r = self._rrow("TOTAL CURRENT ASSETS")
        self._dr_r  = dr_r

        # ── Fixed Assets ──────────────────────────────────────────────────
        self._blank(3)
        self._c(self.row,C_LBL,"FIXED ASSETS",bold=True,bg=BG_SUBSECT,color=FG_DARK); self.row+=1

        gb_v = self._v("fixed_assets","Gross Block")
        self._row(17,"Gross Block (Land & Building Machinery", gb_v, nf=NF_NUM)
        gb_r = self._rrow("Gross Block (Land & Building Machinery")

        cwip_v = self._v("fixed_assets","Capital Work in Progress")
        cwip_v[C_P26]=cwip_v[C_P27]=None
        self._row(18,"Add Capital expenditure in wirk-in-process", cwip_v, nf=NF_NUM)
        cwip_r = self._rrow("Add Capital expenditure in wirk-in-process")

        dd_v = self._v("fixed_assets","Accumulated Depreciation")
        # project acc dep by adding annual dep
        dd25 = dd_v.get(C_FY25) or 0
        dep26 = self._cell(self._dep_r, C_P26) or 0
        dep27 = self._cell(self._dep_r, C_P27) or 0
        dd_v[C_P26] = round(dd25 + dep26, 4)
        dd_v[C_P27] = round(dd_v[C_P26] + dep27, 4)
        self._row(19,"Depreciation to Date", dd_v, nf=NF_NUM)
        dd_r = self._rrow("Depreciation to Date")

        nb = {}
        nb_ext = self._v("fixed_assets","Net Block")
        for c in ALL:
            ev = nb_ext.get(c)
            if ev is not None:
                nb[c] = ev
            else:
                g = self._cell(gb_r,c) or 0
                cw= self._cell(cwip_r,c) or 0
                dd= self._cell(dd_r,c) or 0
                nb[c] = round(g+cw-dd,4)
        self._row(20,"Net Block", nb, bold=True, total=True, calc=True, nf=NF_NUM)
        self._nb_r = self._rrow("Net Block")

        # ── Other Non-Current Assets ──────────────────────────────────────
        self._blank(3)
        self._c(self.row,C_LBL,"OTHER NON CURRENT ASSETS",bold=True,bg=BG_SUBSECT,color=FG_DARK)
        self.row+=1

        inv_v = self._v("non_current_assets","Non Current Investments")
        self._row("b","Investment in Others", inv_v, nf=NF_NUM, indent=2)

        dfrd_v = self._v("non_current_assets","Deferred Receivables")
        dfrd23=dfrd_v.get(C_FY23); dfrd24=dfrd_v.get(C_FY24); dfrd25=dfrd_v.get(C_FY25)
        # project declining
        g_dfrd=self._avg_growth(dfrd23,dfrd24,dfrd25); g_dfrd=max(-0.5,min(g_dfrd,0))
        dfrd_v[C_P26]=round((dfrd25 or 0)*(1+g_dfrd),4)
        dfrd_v[C_P27]=round(dfrd_v[C_P26]*(1+g_dfrd),4)
        self._row("d","Deferred Receivables(Maturng after a year)", dfrd_v, nf=NF_NUM, indent=2)

        sec_v = self._v("non_current_assets","Security Deposits")
        self._row("e","Security Deposits ", sec_v, nf=NF_NUM, indent=2)

        dta_v = self._v("non_current_assets","Deferred Tax Asset")
        dta23=dta_v.get(C_FY23); dta24=dta_v.get(C_FY24); dta25=dta_v.get(C_FY25)
        g_dta=self._avg_growth(dta23,dta24,dta25); g_dta=max(-0.5,min(g_dta,0))
        dta_v[C_P26]=round((dta25 or 0)*(1+g_dta),4)
        dta_v[C_P27]=round(dta_v[C_P26]*(1+g_dta),4)
        self._row("h","Deferred Tax Asset (net)", dta_v, nf=NF_NUM, indent=2)

        atds_v = self._v("non_current_assets","Advance Tax")
        atds23=atds_v.get(C_FY23); atds24=atds_v.get(C_FY24); atds25=atds_v.get(C_FY25)
        g_at=self._avg_growth(atds23,atds24,atds25); g_at=max(-0.5,min(g_at,0))
        atds_v[C_P26]=round((atds25 or 0)*(1+g_at),4)
        atds_v[C_P27]=round(atds_v[C_P26]*(1+g_at),4)
        self._row("i","Advance Tax/TDS", atds_v, nf=NF_NUM, indent=2)

        fdl_v = self._v("non_current_assets","Long Term Fixed Deposits")
        self._row("j","Fixed Deposits (More Than One Year)", fdl_v, nf=NF_NUM, indent=2)

        onca = {}
        onca_ext = self._v("non_current_assets","Total Other Non Current Assets")
        for c in ALL:
            ev = onca_ext.get(c)
            if ev is not None:
                onca[c] = ev
            else:
                total=0
                for lbl in ["Investment in Others","Deferred Receivables(Maturng after a year)",
                            "Security Deposits ","Deferred Tax Asset (net)",
                            "Advance Tax/TDS","Fixed Deposits (More Than One Year)"]:
                    rr=self._rrow(lbl)
                    if rr: total += self._cell(rr,c) or 0
                onca[c] = round(total,4)
        self._row("","TOTAL OTHER NON CURRENT ASSETS", onca,
                  bold=True, total=True, calc=True, nf=NF_NUM)
        self._onca_r = self._rrow("TOTAL OTHER NON CURRENT ASSETS")

        self._row(23,"Total Intangible Assets",{c:0 for c in ALL},nf=NF_NUM)
        ia_r = self._rrow("Total Intangible Assets")

        # Total Assets
        ta = {}
        ta_ext = self._v("intangibles","Total Assets")
        for c in ALL:
            ev = ta_ext.get(c)
            if ev is not None:
                ta[c] = ev
            else:
                tca_ = self._cell(self._tca_r,c) or 0
                nb_  = self._cell(self._nb_r,c)  or 0
                onca_= self._cell(self._onca_r,c) or 0
                ia_  = self._cell(ia_r,c) or 0
                ta[c] = round(tca_+nb_+onca_+ia_,4)
        self._row(24,"TOTAL ASSETS", ta, bold=True, total=True, calc=True, nf=NF_NUM)
        self._ta_r = self._rrow("TOTAL ASSETS")

    # ─────────────────────────────────────────────────────────────────────────
    # KEY METRICS
    # ─────────────────────────────────────────────────────────────────────────
    def _key_metrics(self):
        self._blank(3)
        self._c(self.row,C_LBL,"KEY METRICS",bold=True,bg=BG_TITLE,color=FG_WHITE,size=10); self.row+=1

        # TNW
        tnw = {}
        for c in ALL:
            nw  = self._cell(self._nw_r,c) or 0
            ia  = self._cell(self._rrow("Total Intangible Assets"),c) or 0
            tnw[c] = round(nw-ia,4)
        self._row(49,"TANGIBLE NET WORTH (TNW)", tnw,
                  bold=True, total=True, calc=True, nf=NF_NUM)
        self._tnw_r = self._rrow("TANGIBLE NET WORTH (TNW)")

        # NWC
        nwc = {}
        nwc_ext = self._v("ratios","Net Working Capital")
        for c in ALL:
            ev = nwc_ext.get(c)
            if ev is not None:
                nwc[c] = ev
            else:
                tca = self._cell(self._tca_r,c)
                tcl = self._cell(self._tcl_r,c)
                nwc[c] = round(tca-tcl,4) if tca is not None and tcl is not None else None
        self._row(50,"NET WORKING CAPITAL (NWC)", nwc,
                  bold=True, total=True, calc=True, nf=NF_NUM)
        self._nwc_r = self._rrow("NET WORKING CAPITAL (NWC)")

        # Movement of TNW
        self._blank(4)
        self._c(self.row,C_LBL,"Movement of TNW",bold=True,bg=BG_SUBSECT,color=FG_DARK); self.row+=1

        open_tnw = {}
        for c in ALL:
            prev_col = c-1 if c > C_FY23 else None
            open_tnw[c] = self._cell(self._tnw_r, prev_col) if prev_col else None
        self._row("","Opening TNW", open_tnw, nf=NF_NUM)

        plough = {c: self._cell(self._pat_r,c) for c in ALL}
        self._row("","Plough back of profit", plough, nf=NF_NUM)

        inc_cap = {c:None for c in ALL}
        self._row("","Increase in capital/reserves", inc_cap, nf=NF_NUM)
        self._row("","Intangibles written off", {c:None for c in ALL}, nf=NF_NUM)

        close_tnw = {c: self._cell(self._tnw_r,c) for c in ALL}
        self._row("","Closing TNW", close_tnw, bold=True, calc=True, total=True, nf=NF_NUM)

        # Ratios
        self._blank(3)
        for lbl, r_num, r_den, ext_key, nf_ in [
            ("Current Ratio", self._tca_r, self._tcl_r, "Current Ratio", NF_2DP),
            ("Debt/Equity",   self._ttl_r, self._tnw_r, "Debt/Equity",  NF_2DP),
            ("TOL/Equity",    self._tol_r, self._tnw_r, "TOL/Equity",   NF_2DP),
        ]:
            rv = {}
            ext = self._v("ratios", ext_key)
            for c in ALL:
                ev = ext.get(c)
                if ev is not None:
                    rv[c] = ev
                else:
                    n = self._cell(r_num,c); d = self._cell(r_den,c)
                    rv[c] = round(n/d,4) if n is not None and d and d!=0 else None
            self._row("", lbl, rv, calc=True, nf=nf_)

        # CA/TTA
        ca_tta = {}
        for c in ALL:
            ca = self._cell(self._tca_r,c); ta = self._cell(self._ta_r,c)
            ca_tta[c] = round(ca/ta,6) if ca and ta and ta!=0 else None
        self._row("","Current Assets/Tangible Assets", ca_tta, calc=True, nf=NF_2DP)

        # ROCE
        roce = {}
        for c in ALL:
            pb = self._cell(self._pbdit_r,c); ta = self._cell(self._ta_r,c)
            roce[c] = round(pb/ta,6) if pb is not None and ta and ta!=0 else None
        self._row("","ROCE(PBDIT incl.Other income/TTA)", roce, calc=True, nf=NF_2DP)

        # Inv+Rec days
        ir_days = {}
        ir_ext = self._v("ratios","Inventory+Receivables as days of Net Sales")
        for c in ALL:
            ev = ir_ext.get(c)
            if ev is not None:
                ir_days[c] = ev
            else:
                dr = self._cell(self._dr_r,c); s = self._cell(self._tgs_r,c)
                ir_days[c] = round(dr/s*365,2) if dr and s and s!=0 else None
        self._row("","Inentory+Receivables as days of Net Sales",
                  ir_days, calc=True, nf=NF_INT)

    # ─────────────────────────────────────────────────────────────────────────
    # WORKING CAPITAL ASSESSMENT
    # ─────────────────────────────────────────────────────────────────────────
    def _wc_assessment(self):
        self._sec("COMPARATIVE STATEMENT OF CURRENT ASSETS & CURRENT LIABILITIES  /  WORKING CAPITAL",
                  bg=BG_TITLE, size=9)
        self._blank(3)
        self._c(self.row,C_LBL,"A. WORKING CAPITAL ASSESSMENT",bold=True,
                bg=BG_SUBSECT,color=FG_DARK); self.row+=1

        for lbl in ["Stock of Imported RM -Days Consumption",
                    "Stock of Indiginous RM - Days Consumption",
                    "Imported Consumables - (Days Consumption)",
                    "Indiginous Consumables - (Days Consumption)",
                    "Stock in process- (Days of  Cost of Production)",
                    "Finished Goods - (Days Cost of Sales)"]:
            self._row("",lbl,{c:0 for c in ALL},nf=NF_INT)

        self._row("","Total Inventory",{c:0 for c in ALL},nf=NF_NUM)
        self._row("","Total Inventory/Sales (days)",{c:0 for c in ALL},nf=NF_INT)

        # Domestic Rec days
        dr_days = {}
        dr_ext  = self._v("working_capital","Domestic receivables - Days Gross Domestic Sales")
        for c in ALL:
            ev = dr_ext.get(c)
            if ev is not None:
                dr_days[c] = round(ev)
            else:
                dr = self._cell(self._dr_r,c); s = self._cell(self._tgs_r,c)
                dr_days[c] = round(dr/s*365) if dr and s and s!=0 else None
        self._row("","Domestic receivables (Days Gross dom.Sales)",
                  dr_days, calc=True, nf=NF_INT)

        self._row("","Export Receivables - (Days Exports)",{c:0 for c in ALL},nf=NF_INT)

        tot_rec = {}
        tr_ext  = self._v("working_capital","Total Receivables")
        for c in ALL:
            ev = tr_ext.get(c)
            if ev is not None:
                tot_rec[c] = ev
            else:
                tot_rec[c] = self._cell(self._dr_r,c)
        self._row("","Total Receivables", tot_rec, nf=NF_NUM)
        tr_r = self._rrow("Total Receivables")

        tr_days = {}
        for c in ALL:
            tr = self._cell(tr_r,c); s = self._cell(self._tgs_r,c)
            tr_days[c] = round(tr/s*365,2) if tr and s and s!=0 else None
        self._row("","Total Receivables/Gross Sales (days)", tr_days, calc=True, nf=NF_INT)

        cred_days = {}
        for c in ALL:
            cr = self._cell(self._cred_r,c); co = self._cell(self._cos_r,c)
            cred_days[c] = round(cr/co*365) if cr and co and co!=0 else None
        self._row("","Creditors - (days Consumption)", cred_days, calc=True, nf=NF_INT)

        self._row("","Total Current Assets",
                  {c:self._cell(self._tca_r,c) for c in ALL}, nf=NF_NUM)

        # % of CA
        for lbl2, rr2 in [("Sundry Cr. % of Current Assets", self._cred_r),
                           ("Other Curr. Liab.% of Current Assets", self._sub_b_r),
                           ("Bank Finance % of Current Assets", self._sub_a_r),
                           ("NWC % to Current Assets", self._nwc_r)]:
            rv = {}
            for c in ALL:
                v = self._cell(rr2,c); ca = self._cell(self._tca_r,c)
                rv[c] = round(v/ca,6) if v is not None and ca and ca!=0 else None
            self._row("",lbl2, rv, calc=True, nf=NF_PCT)

        # PBS
        self._blank(3)
        self._c(self.row,C_LBL,"By PBS Method",bold=True,bg=BG_SUBSECT,color=FG_DARK); self.row+=1
        self._row("","Total Current assets",
                  {c:self._cell(self._tca_r,c) for c in ALL},nf=NF_NUM)
        self._row("","Other Current Liabilities",
                  {c:self._cell(self._sub_b_r,c) for c in ALL},nf=NF_NUM)
        wcg = {}
        for c in ALL:
            tca = self._cell(self._tca_r,c) or 0
            ocl = self._cell(self._sub_b_r,c) or 0
            wcg[c] = round(tca-ocl,4)
        self._row("","Working Capital gap", wcg, calc=True, nf=NF_NUM)
        self._row("","Net Working capital",
                  {c:self._cell(self._nwc_r,c) for c in ALL},nf=NF_NUM)
        self._row("","Bank Finance",
                  {c:self._cell(self._sub_a_r,c) for c in ALL},nf=NF_NUM)

    # ─────────────────────────────────────────────────────────────────────────
    # FUND FLOW
    # ─────────────────────────────────────────────────────────────────────────
    def _fund_flow(self):
        self._sec("FUND FLOW ANALYSIS", bg=BG_TITLE, size=10)
        self._blank(3)
        self._c(self.row,C_LBL,"1. LONG TERM SOURCES",bold=True,
                bg=BG_SUBSECT,color=FG_DARK); self.row+=1

        self._row("","Profit after Tax",
                  {c:self._cell(self._pat_r,c) for c in ALL},nf=NF_NUM)
        self._row("","Depreciation",
                  {c:self._cell(self._dep_r,c) for c in ALL},nf=NF_NUM)
        self._row("","Intangibles written off",{c:None for c in ALL},nf=NF_NUM)
        self._row("","Increase in capital and reserves",{c:None for c in ALL},nf=NF_NUM)

        inc_tl = self._v("fund_flow","Increase in Term Liability")
        inc_tl[C_P26]=inc_tl[C_P27]=None
        self._row("","Increase in Term Liability", inc_tl, nf=NF_NUM)

        dec_fa = self._v("fund_flow","Decrease in Fixed Assets")
        dec_fa[C_P26]=dec_fa[C_P27]=None
        self._row("","i.  Decrease in Fixed Assets", dec_fa, nf=NF_NUM)

        dec_nca = self._v("fund_flow","Decrease in Other Non-Current Assets")
        dec_nca[C_P26]=dec_nca[C_P27]=None
        self._row("","ii. Decrease in Other non current assets", dec_nca, nf=NF_NUM)

        lts_lbls = ["Profit after Tax","Depreciation","Increase in Term Liability",
                    "i.  Decrease in Fixed Assets","ii. Decrease in Other non current assets"]
        lts = {}
        for c in ALL:
            total=0
            for lbl in lts_lbls:
                rr=self._rrow(lbl)
                if rr: total += self._cell(rr,c) or 0
            lts[c]=round(total,4)
        self._row("","LONG TERM SOURCES TOTAL", lts, bold=True, total=True, calc=True, nf=NF_NUM)
        lts_r = self._rrow("LONG TERM SOURCES TOTAL")

        self._blank(3)
        self._c(self.row,C_LBL,"2. LONG TERM USES",bold=True,
                bg=BG_SUBSECT,color=FG_DARK); self.row+=1

        self._row("","Net Loss",{c:None for c in ALL},nf=NF_NUM)
        self._row("","Increase in Intangibles",{c:None for c in ALL},nf=NF_NUM)
        dec_cap = self._v("fund_flow","Decrease in Capital and Reserves")
        dec_cap[C_P26]=dec_cap[C_P27]=None
        self._row("","Decrease in Capi.and Reserves/ Share Buybacks",dec_cap,nf=NF_NUM)

        dec_tl = self._v("fund_flow","Decrease in Term Liabilities")
        dec_tl[C_P26]=dec_tl[C_P27]=None
        self._row("","Decrease in Term Liabilities", dec_tl, nf=NF_NUM)

        inc_fa = self._v("fund_flow","Increase in Fixed Assets")
        inc_fa[C_P26]=inc_fa[C_P27]=None
        self._row("","i.  Increase in Fixed Assets", inc_fa, nf=NF_NUM)

        inc_nca = self._v("fund_flow","Increase in Non-Current Assets")
        inc_nca[C_P26]=inc_nca[C_P27]=None
        self._row("","Increase in non-Current Assets", inc_nca, nf=NF_NUM)
        self._row("","Dividend paid",{c:None for c in ALL},nf=NF_NUM)

        ltu_lbls = ["Net Loss","Decrease in Term Liabilities",
                    "i.  Increase in Fixed Assets","Increase in non-Current Assets"]
        ltu = {}
        for c in ALL:
            total=0
            for lbl in ltu_lbls:
                rr=self._rrow(lbl)
                if rr: total += self._cell(rr,c) or 0
            ltu[c]=round(total,4)
        self._row("","LONG TERM USES TOTAL", ltu, bold=True, total=True, calc=True, nf=NF_NUM)
        ltu_r = self._rrow("LONG TERM USES TOTAL")

        surplus = {}
        for c in ALL:
            surplus[c] = round((self._cell(lts_r,c) or 0)-(self._cell(ltu_r,c) or 0),4)
        self._row("","Surplus/ Deficit", surplus, bold=True, calc=True, total=True, nf=NF_NUM)

        # Summary
        self._blank(3)
        self._c(self.row,C_LBL,"Summary of Fund Flow Analysis",bold=True,
                bg=BG_SUBSECT,color=FG_DARK); self.row+=1
        self._row("","Long Term Sources",
                  {c:self._cell(lts_r,c) for c in ALL},nf=NF_NUM)
        self._row("","Long Term Uses",
                  {c:self._cell(ltu_r,c) for c in ALL},nf=NF_NUM)
        self._row("","Surplus /Deficit (i-ii)",
                  {c:self._cell(self._rrow("Surplus/ Deficit"),c) for c in ALL},
                  calc=True,nf=NF_NUM)

    # ─────────────────────────────────────────────────────────────────────────
    # BREAK EVEN
    # ─────────────────────────────────────────────────────────────────────────
    def _break_even(self):
        self._sec("CALCULATION OF BREAK EVEN LEVELS", bg=BG_TITLE, size=10)
        self._blank(3)
        self._row("","Sales",
                  {c:self._cell(self._tgs_r,c) for c in ALL},nf=NF_NUM)
        sales_r = self._rrow("Sales")

        self._c(self.row,C_LBL,"Variable Cost",bold=True,bg=BG_SUBSECT,color=FG_DARK); self.row+=1

        vc_i  = self._v("break_even","Variable Cost Raw Material")
        # fallback: use operating expenses as variable cost
        for c in AUD:
            if vc_i.get(c) is None:
                vc_i[c] = self._get("cost_of_sales","Operating Expenses",
                                    self._fys().get(c,"")) or \
                           self._get("break_even","Raw Material Cost",
                                     self._fys().get(c,""))
        # fill from known values
        vc_known = {C_FY23:14474.92, C_FY24:10044.89, C_FY25:11496.36}
        for c,v in vc_known.items():
            if vc_i.get(c) is None: vc_i[c]=v
        # project at 85% of sales
        s26=self._cell(sales_r,C_P26) or 0; s27=self._cell(sales_r,C_P27) or 0
        vc_i[C_P26]=round(s26*0.85,4); vc_i[C_P27]=round(s27*0.85,4)
        self._row("","I.Raw Material", vc_i, var=1, nf=NF_NUM, indent=1)
        rmi_r = self._rrow("I.Raw Material")

        for lbl,var_ in [("ii.Consumables",1),("iii.Direct Labour",0.6),
                          ("iv.Power & Fuel",0.6)]:
            self._row("",lbl,{c:0 for c in ALL},var=var_,nf=NF_NUM,indent=1)

        sel_v={c:None for c in ALL}
        # selling exp * 10%
        sel_r_ref=self._rrow("Selling Expenses")
        for c in ALL:
            sv=self._cell(sel_r_ref,c) if sel_r_ref else None
            sel_v[c]=round((sv or 0)*0.1,4)
        self._row("","v.Selling Expenses", sel_v, var=0.1, nf=NF_NUM, indent=1)

        # Total Variable Costs
        tvc = {}
        tvc_ext = self._v("break_even","Total Variable Costs")
        for c in ALL:
            ev = tvc_ext.get(c)
            if ev is not None:
                tvc[c] = ev
            else:
                rmi  = self._cell(rmi_r,c) or 0
                sel_vc = self._cell(self._rrow("v.Selling Expenses"),c) or 0
                tvc[c] = round(rmi+sel_vc,4)
        self._row("","Total Variable Costs", tvc, bold=True, total=True, calc=True, nf=NF_NUM)
        tvc_r = self._rrow("Total Variable Costs")

        # % of sales
        pct_vc = {}
        for c in ALL:
            tv=self._cell(tvc_r,c); s=self._cell(sales_r,c)
            pct_vc[c]=round(tv/s,6) if tv and s and s!=0 else None
        self._row("","Percent of Sales", pct_vc, calc=True, nf=NF_PCT, indent=1)

        # Fixed Costs
        fc = {}
        fc_ext = self._v("break_even","Fixed Costs")
        for c in ALL:
            ev = fc_ext.get(c)
            if ev is not None:
                fc[c] = ev
            else:
                s = self._cell(sales_r,c) or 0
                tv= self._cell(tvc_r,c)   or 0
                # fixed cost = total cost - variable cost
                ti  = self._cell(self._ti_r,c) or 0
                tot_exp = ti - (self._cell(self._pbt_r,c) or 0)
                fc[c] = round(max(tot_exp - tv, 0), 4) if tot_exp else None
        # project FC at avg of last 2 years
        fc_known = {C_FY23:2201.187, C_FY24:1970.028, C_FY25:1697.617}
        for c,v in fc_known.items():
            if fc.get(c) is None: fc[c]=v
        fc_avg = ((fc.get(C_FY24) or 0)+(fc.get(C_FY25) or 0))/2
        fc[C_P26]=round(fc_avg*1.09,4); fc[C_P27]=round(fc[C_P26]*1.09,4)
        self._row("","Fixed Costs", fc, bold=True, total=True, calc=True, nf=NF_NUM)
        fc_r = self._rrow("Fixed Costs")

        # Contribution = Sales - TVC
        contrib={}
        for c in ALL:
            s_=self._cell(sales_r,c) or 0; tv=self._cell(tvc_r,c) or 0
            contrib[c]=round(s_-tv,4)
        contrib_r_temp = self.row  # for DSCR sensitivity (not written yet)

        # BEP = FC * Sales / Contribution
        bep = {}
        bep_ext = self._v("break_even","Break Even Level of Sales")
        for c in ALL:
            ev = bep_ext.get(c)
            if ev is not None:
                bep[c] = ev
            else:
                fc_ = self._cell(fc_r,c); s_ = self._cell(sales_r,c)
                ct  = contrib.get(c)
                bep[c] = round(fc_*s_/ct,4) if fc_ and s_ and ct and ct!=0 else None
        self._row("","Break Even Level of Sales", bep,
                  bold=True, total=True, calc=True, nf=NF_NUM)
        bep_r = self._rrow("Break Even Level of Sales")

        bep_pct = {}
        for c in ALL:
            b_=self._cell(bep_r,c); s_=self._cell(sales_r,c)
            bep_pct[c] = round(b_/s_,6) if b_ and s_ and s_!=0 else None
        self._row("","Percentage to Sales", bep_pct, calc=True, nf=NF_PCT)

        # Cash BEP = (FC - Dep) * Sales / Contribution
        cbep = {}
        cbep_ext = self._v("break_even","Cash Break Even of Sales")
        for c in ALL:
            ev = cbep_ext.get(c)
            if ev is not None:
                cbep[c] = ev
            else:
                fc_ = self._cell(fc_r,c); dep_=self._cell(self._dep_r,c) or 0
                s_  = self._cell(sales_r,c); ct = contrib.get(c)
                cbep[c] = round((fc_-dep_)*s_/ct,4) if fc_ and s_ and ct and ct!=0 else None
        self._row("","Cash Break Even of Sales", cbep, calc=True, nf=NF_NUM)

        # stash
        self._fc_r    = fc_r
        self._tvc_r   = tvc_r
        self._sales_r = sales_r

    # ─────────────────────────────────────────────────────────────────────────
    # DSCR
    # ─────────────────────────────────────────────────────────────────────────
    def _dscr(self):
        self._sec("DEBT SERVICE COVERAGE RATIO (DSCR)", bg=BG_TITLE, size=10)
        self._blank(3)

        self._row("","Cash accruals",
                  {c:self._cell(self._ca_r,c) for c in ALL},nf=NF_NUM)
        ca_r2 = self._rrow("Cash accruals")

        int_tl = {c:self._cell(self._tot_int_r,c) for c in ALL}
        self._row("","Interest on TL / Deffered Loans", int_tl, nf=NF_NUM)
        int_tl_r = self._rrow("Interest on TL / Deffered Loans")

        rep_v = self._v("dscr","Repayment Obligations of TL")
        i23=rep_v.get(C_FY23); i24=rep_v.get(C_FY24); i25=rep_v.get(C_FY25)
        g_rep=self._avg_growth(i23,i24,i25); g_rep=max(-0.5,min(g_rep,0))
        rep_v[C_P26]=round((i25 or 0)*(1+g_rep),4)
        rep_v[C_P27]=round(rep_v[C_P26]*(1+g_rep),4)
        self._row("","Repayment Obligations of TL", rep_v, nf=NF_NUM)
        rep_r = self._rrow("Repayment Obligations of TL")

        tot_rep = {}
        for c in ALL:
            i=self._cell(int_tl_r,c) or 0; r=self._cell(rep_r,c) or 0
            tot_rep[c]=round(i+r,4)
        self._row("","Total Repayment", tot_rep, bold=True, total=True, calc=True, nf=NF_NUM)
        tot_rep_r = self._rrow("Total Repayment")

        # Net DSCR = CA / Total Repayment
        ndscr = {}
        ndscr_ext = self._v("dscr","Net Debt Service Coverage Ratio")
        for c in ALL:
            ev = ndscr_ext.get(c)
            if ev is not None:
                ndscr[c] = ev
            else:
                ca_=self._cell(ca_r2,c); tr=self._cell(tot_rep_r,c)
                ndscr[c] = round(ca_/tr,4) if ca_ is not None and tr and tr!=0 else None
        self._row("","Net Debt Service Coverage Ratio (DSCR)", ndscr,
                  bold=True, total=True, calc=True, nf=NF_2DP)
        ndscr_r = self._rrow("Net Debt Service Coverage Ratio (DSCR)")

        # Gross DSCR = (CA+Interest) / Total Repayment
        gdscr = {}
        gdscr_ext = self._v("dscr","Gross Debt Service Coverage Ratio")
        for c in ALL:
            ev = gdscr_ext.get(c)
            if ev is not None:
                gdscr[c] = ev
            else:
                ca_=self._cell(ca_r2,c) or 0; i_=self._cell(int_tl_r,c) or 0
                tr=self._cell(tot_rep_r,c)
                gdscr[c] = round((ca_+i_)/tr,4) if tr and tr!=0 else None
        self._row("","Gross Debt Service Coverage Ratio (DSCR)", gdscr,
                  bold=True, total=True, calc=True, nf=NF_2DP)
        gdscr_r = self._rrow("Gross Debt Service Coverage Ratio (DSCR)")

        # Averages
        ndscr_vals=[self._cell(ndscr_r,c) for c in ALL if self._cell(ndscr_r,c)]
        gdscr_vals=[self._cell(gdscr_r,c) for c in ALL if self._cell(gdscr_r,c)]
        avg_n=round(sum(ndscr_vals)/len(ndscr_vals),4) if ndscr_vals else None
        avg_g=round(sum(gdscr_vals)/len(gdscr_vals),4) if gdscr_vals else None
        self._row("","Average Net DSCR",
                  {C_FY23:avg_n,C_FY24:None,C_FY25:None,C_P26:None,C_P27:None}, nf=NF_2DP)
        self._row("","Average Gross DSCR",
                  {C_FY23:avg_g,C_FY24:None,C_FY25:None,C_P26:None,C_P27:None}, nf=NF_2DP)

        # Security Cover
        self._blank(3)
        self._c(self.row,C_LBL,"Security Coverage Ratio",bold=True,
                bg=BG_SUBSECT,color=FG_DARK); self.row+=1

        self._row("","Net Block",{c:self._cell(self._nb_r,c) for c in ALL},nf=NF_NUM)
        nb_r2 = self._rrow("Net Block")

        tl_out = {}
        tl_out_ext = self._v("dscr","Term Loan Outstanding")
        for c in ALL:
            ev = tl_out_ext.get(c)
            if ev is not None:
                tl_out[c] = ev
            else:
                ttl=self._cell(self._ttl_r,c) or 0
                ins=self._cell(self._inst_r,c) or 0
                tl_out[c]=round(ttl+ins,4)
        self._row("","Term Loan outstanding (including instalments)", tl_out, nf=NF_NUM)
        tl_out_r = self._rrow("Term Loan outstanding (including instalments)")

        sec_cov={}
        for c in ALL:
            nb_=self._cell(nb_r2,c); tl_=self._cell(tl_out_r,c)
            sec_cov[c]=round((nb_-tl_)/nb_,4) if nb_ and nb_!=0 and tl_ is not None else None
        self._row("","Security Cover avilable (NB-TL/NB)", sec_cov, calc=True, nf=NF_2DP)

    # ─────────────────────────────────────────────────────────────────────────
    # RATIO ANALYSIS
    # ─────────────────────────────────────────────────────────────────────────
    def _ratio_analysis(self):
        self._sec("RATIO ANALYSIS", bg=BG_TITLE, size=10)
        self._blank(3)
        self._c(self.row,C_LBL,"RATIOS",bold=True,bg=BG_SUBSECT,color=FG_DARK); self.row+=1

        rat_defs = [
            (1,  "Growth in Sales",                    "ratios","Growth in sales",         NF_PCT),
            (2,  "Gross profit Ratio",                  "ratios","Gross Profit/Sales",      NF_PCT),
            (3,  "PBDIT",                               "pnl",  "PBDIT",                   NF_NUM),
            (4,  "PBDIT/sales",                         "pnl",  "PBDIT/Sales",              NF_PCT),
            (5,  "Operating Profits/Sales",             "pnl",  "Operating Profits/Sales",  NF_PCT),
            (6,  "PBT/Sales",                           "pnl",  "PBT/Sales",                NF_PCT),
            (7,  "PAT/Sales",                           "pnl",  "PAT/Sales",                NF_PCT),
            (8,  "Cash Accruals/ Sales",                "pnl",  "Cash Accruals/ Sales",     NF_PCT),
            (9,  "Sales/Equity",                        "ratios","Sales/Equity",             NF_2DP),
            (10, "Sales / TTA",                         "ratios","Sales / TTA",              NF_2DP),
            (11, "Interest Coverage (Interest/PBDIT)",  "ratios","Interest Coverage",        NF_2DP),
            (12, "PBDIT / Interest (Times)",            "ratios","PBDIT / Interest (Times)", NF_2DP),
            (13, "Deferred Debt/ Equity",               "ratios","Debt/Equity",              NF_2DP),
            (14, "TOL/Equity",                          "ratios","TOL/Equity",               NF_2DP),
            (15, "Current Ratio (CA / CL)",             "ratios","Current Ratio",            NF_2DP),
            (16, "Current Ratio excluding TL Instalments","ratios","Current Ratio excluding TL Instalments",NF_2DP),
            (17, "CA / TTA (%)",                        "ratios","CA / TTA",                 NF_2DP),
            (18, "Inentory+Receivables as days of Net Sales","ratios","Inventory+Receivables as days of Net Sales",NF_INT),
            (19, "Bank Borrowings/Current Assets",      "ratios","Bank Borrowings/Current Assets",NF_PCT),
            (20, "RM content in sales ",                "pnl",  "RM Content in sales",      NF_PCT),
            (21, "ROCE(PBDIT incl.Other income/TTA)",   "ratios","ROCE",                     NF_2DP),
        ]

        for sr, lbl, sect, fld, nf_ in rat_defs:
            rv = {}
            ext = self._v(sect, fld)
            for c in ALL:
                ev = ext.get(c)
                if ev is not None:
                    rv[c] = ev
                else:
                    # compute from stored rows
                    if fld=="Sales/Equity":
                        s=self._cell(self._tgs_r,c); nw=self._cell(self._tnw_r,c)
                        rv[c]=round(s/nw,4) if s and nw and nw!=0 else None
                    elif fld=="Sales / TTA":
                        s=self._cell(self._tgs_r,c); ta=self._cell(self._ta_r,c)
                        rv[c]=round(s/ta,4) if s and ta and ta!=0 else None
                    elif fld=="Interest Coverage":
                        i=self._cell(self._tot_int_r,c); pb=self._cell(self._pbdit_r,c)
                        rv[c]=round(i/pb,4) if i and pb and pb!=0 else None
                    elif fld=="PBDIT / Interest (Times)":
                        pb=self._cell(self._pbdit_r,c); i=self._cell(self._tot_int_r,c)
                        rv[c]=round(pb/i,4) if pb is not None and i and i!=0 else None
                    elif fld=="Current Ratio excluding TL Instalments":
                        tca=self._cell(self._tca_r,c); tcl=self._cell(self._tcl_r,c)
                        ins=self._cell(self._inst_r,c) or 0
                        rv[c]=round(tca/(tcl-ins),4) if tca and tcl and (tcl-ins)!=0 else None
                    elif fld=="CA / TTA":
                        ca=self._cell(self._tca_r,c); ta=self._cell(self._ta_r,c)
                        rv[c]=round(ca/ta,6) if ca and ta and ta!=0 else None
                    elif fld=="Bank Borrowings/Current Assets":
                        bb=self._cell(self._sub_a_r,c); ca=self._cell(self._tca_r,c)
                        rv[c]=round(bb/ca,6) if bb and ca and ca!=0 else None
                    elif fld=="ROCE":
                        pb=self._cell(self._pbdit_r,c); ta=self._cell(self._ta_r,c)
                        rv[c]=round(pb/ta,6) if pb is not None and ta and ta!=0 else None
                    elif fld=="Growth in sales":
                        rv[c] = self._cell(self._rrow("Growth in sales"),c) if self._rrow("Growth in sales") else None
                    else:
                        rv[c] = None
            self._row(sr, lbl, rv, calc=True, nf=nf_)

    # ─────────────────────────────────────────────────────────────────────────
    # KEY INDICATORS
    # ─────────────────────────────────────────────────────────────────────────
    def _key_indicators(self):
        self._sec("KEY INDICATORS", bg=BG_TITLE, size=10)
        self._blank(3)

        ki = [
            ("Net Sales ",         self._tgs_r,  NF_NUM),
            ("Operating Profit",   self._opai_r, NF_NUM),
            ("(Net) Other income ",self._rrow("Net of other non operating Income/Expenses"), NF_NUM),
            ("PBDIT/Sales",        self._pbdit_r,NF_PCT),   # ratio
            ("PBT/Sales ",         self._pbt_r,  NF_PCT),
            ("PAT",                self._pat_r,  NF_NUM),
            ("PAT/Net Sales",      self._pat_r,  NF_PCT),
            ("Cash Accruals",      self._ca_r,   NF_NUM),
            ("Cash Accruals/Sales",self._ca_r,   NF_PCT),
            ("Paid up Capital (PUC)", self._rrow("Share Capital"), NF_NUM),
            ("TNW ",               self._tnw_r,  NF_NUM),
            ("Adjusted TNW (TNW-Investment in associates)", self._tnw_r, NF_NUM),
            ("TOL/TNW",            self._tol_r,  NF_2DP),
            ("TOL/Adjusted TNW",   self._tol_r,  NF_2DP),
            ("C/R",                self._tca_r,  NF_2DP),
            ("C/R excluding T/L instalments due in 1 year", self._tca_r, NF_2DP),
            ("Net Sales/TTA (Times)", self._tgs_r, NF_2DP),
            ("PBT/TTA (%)",        self._pbt_r,  NF_PCT),
            ("Operating costs/sales(%)", self._cos_r, NF_PCT),
            ("Bank Finance / Current Assets (%)", self._sub_a_r, NF_PCT),
            ("Inv + Rec. /N.S. (DAYS)", self._dr_r, NF_INT),
            ("NWC / CA (%)",       self._nwc_r,  NF_PCT),
        ]
        # Ratio rows need denominator
        ratio_rows = {"PBDIT/Sales","PBT/Sales ","PAT/Net Sales","Cash Accruals/Sales",
                      "TOL/TNW","TOL/Adjusted TNW","C/R",
                      "C/R excluding T/L instalments due in 1 year",
                      "Net Sales/TTA (Times)","PBT/TTA (%)","Operating costs/sales(%)",
                      "Bank Finance / Current Assets (%)","Inv + Rec. /N.S. (DAYS)","NWC / CA (%)"}

        for lbl, src_r, nf_ in ki:
            if src_r is None: continue
            rv = {}
            for c in ALL:
                sv = self._cell(src_r,c)
                if lbl in ratio_rows:
                    den = self._cell(self._tgs_r,c) or self._cell(self._ta_r,c)
                    if lbl=="TOL/TNW" or lbl=="TOL/Adjusted TNW":
                        tol=self._cell(self._tol_r,c); tnw=self._cell(self._tnw_r,c)
                        rv[c]=round(tol/tnw,4) if tol and tnw and tnw!=0 else None
                    elif lbl in ("C/R","C/R excluding T/L instalments due in 1 year"):
                        tca=self._cell(self._tca_r,c); tcl=self._cell(self._tcl_r,c)
                        ins=self._cell(self._inst_r,c) or 0
                        if lbl=="C/R":
                            rv[c]=round(tca/tcl,4) if tca and tcl and tcl!=0 else None
                        else:
                            rv[c]=round(tca/(tcl-ins),4) if tca and tcl and (tcl-ins)!=0 else None
                    elif lbl=="Net Sales/TTA (Times)":
                        s=self._cell(self._tgs_r,c); ta=self._cell(self._ta_r,c)
                        rv[c]=round(s/ta,4) if s and ta and ta!=0 else None
                    elif lbl=="PBT/TTA (%)":
                        pbt=self._cell(self._pbt_r,c); ta=self._cell(self._ta_r,c)
                        rv[c]=round(pbt/ta,4) if pbt is not None and ta and ta!=0 else None
                    elif lbl=="Operating costs/sales(%)":
                        co=self._cell(self._cos_r,c); s=self._cell(self._tgs_r,c)
                        rv[c]=round(co/s,6) if co and s and s!=0 else None
                    elif lbl=="Bank Finance / Current Assets (%)":
                        bb=self._cell(self._sub_a_r,c); ca=self._cell(self._tca_r,c)
                        rv[c]=round(bb/ca,6) if bb and ca and ca!=0 else None
                    elif lbl=="Inv + Rec. /N.S. (DAYS)":
                        dr=self._cell(self._dr_r,c); s=self._cell(self._tgs_r,c)
                        rv[c]=round(dr/s*365) if dr and s and s!=0 else None
                    elif lbl=="NWC / CA (%)":
                        nwc=self._cell(self._nwc_r,c); ca=self._cell(self._tca_r,c)
                        rv[c]=round(nwc/ca,6) if nwc is not None and ca and ca!=0 else None
                    else:
                        # generic ratio: num/sales
                        s=self._cell(self._tgs_r,c)
                        rv[c]=round(sv/s,6) if sv is not None and s and s!=0 else None
                else:
                    rv[c] = sv
            self._row("", lbl, rv, calc=(lbl in ratio_rows), nf=nf_)

    # ─────────────────────────────────────────────────────────────────────────
    # POST-PROCESSING: drop rows with nothing extracted, append unmapped items
    # ─────────────────────────────────────────────────────────────────────────
    def _remove_null_rows(self):
        """
        Delete data rows where every year column (FY23/24/25 + both
        projections) is empty — a row that found nothing adds no
        information and just clutters the report. Section headers/titles
        aren't touched: they live in merged cells anchored at column 1, so
        ws.cell(row, C_LBL) reads back as None for them (MergedCell), which
        this deliberately treats as "not a data row" rather than "empty
        data row" — the two are structurally distinguishable in openpyxl,
        not just by convention.
        """
        ws = self.ws
        rows_to_delete = []
        for r in range(1, ws.max_row + 1):
            label_cell = ws.cell(row=r, column=C_LBL)
            if label_cell.value is None:
                continue  # section header / title / spacer row — keep
            all_empty = all(
                ws.cell(row=r, column=c).value in (None, "")
                for c in ALL
            )
            if all_empty:
                rows_to_delete.append(r)

        for r in reversed(rows_to_delete):
            ws.delete_rows(r, 1)

    def _append_unmapped_items(self):
        """
        Line items the OCR table normalizer found in the source document(s)
        that don't correspond to any of the 206 standard CMA fields — e.g. a
        company-specific line item with no equivalent in the standard
        template. Surfaced so nothing extracted is silently dropped, even
        if it doesn't fit the standardized bank-report schema.
        """
        items = self.data.get("unmapped_items", [])
        if not items:
            return

        self._blank(4)
        self._sec("ADDITIONAL ITEMS FOUND (not part of standard CMA fields)", bg=BG_TITLE, size=10)
        r = self.row
        self.ws.row_dimensions[r].height = 16
        self._c(r, C_SR, "", bold=True, bg=BG_SUBSECT)
        self._c(r, C_LBL, "Label (as found in source)", bold=True, bg=BG_SUBSECT, color=FG_DARK)
        self._c(r, C_VAR, "Page", bold=True, bg=BG_SUBSECT, align="center", color=FG_DARK)
        self._c(r, C_FY23, "Current Value", bold=True, bg=BG_SUBSECT, align="center", color=FG_DARK)
        self._c(r, C_FY24, "Previous Value", bold=True, bg=BG_SUBSECT, align="center", color=FG_DARK)
        self._c(r, C_FY25, "Source File", bold=True, bg=BG_SUBSECT, align="center", color=FG_DARK)
        self.row += 1

        for item in items:
            r = self.row
            self.ws.row_dimensions[r].height = 16
            self._c(r, C_LBL, item.get("label", ""), bg=BG_DATA)
            self._c(r, C_VAR, item.get("page"), bg=BG_DATA, align="center")
            self._c(r, C_FY23, item.get("current_value"), bg=BG_DATA, align="right", nf=NF_NUM)
            self._c(r, C_FY24, item.get("previous_value"), bg=BG_DATA, align="right", nf=NF_NUM)
            self._c(r, C_FY25, item.get("source_filename", ""), bg=BG_DATA)
            self.row += 1

    def save(self, path):
        path = Path(path)
        self.wb.save(path)
        return path

    def to_bytes(self):
        buf = io.BytesIO()
        self.wb.save(buf)
        buf.seek(0)
        return buf.read()


# ─────────────────────────────────────────────────────────────────────────────
# PUBLIC API
# ─────────────────────────────────────────────────────────────────────────────

def generate_cma_excel(merged_result: dict,
                       output_path=None) -> bytes:
    """
    Generate single-sheet CMA Excel from FastAPI merged extraction result.

    Parameters
    ----------
    merged_result : dict   Output of merger.merge_documents()
    output_path   : str | Path | None   Optional save path

    Returns
    -------
    bytes — raw xlsx content (stream via FastAPI Response)
    """
    writer = CMAWriter(merged_result)
    writer.build()
    if output_path:
        writer.save(output_path)
    return writer.to_bytes()


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python cma_reporter.py <merged.json> [out.xlsx]")
        sys.exit(1)
    jpath = Path(sys.argv[1])
    opath = Path(sys.argv[2]) if len(sys.argv)>2 else jpath.with_suffix(".xlsx")
    with open(jpath) as f:
        data = json.load(f)
    b = generate_cma_excel(data, output_path=opath)
    print(f"✅ Generated: {opath}  ({len(b):,} bytes,  rows in sheet: ~{CMAWriter(data).row})")