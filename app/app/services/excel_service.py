from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.exceptions import ExcelGenerationFailureError
from app.schemas.common import DocumentType, ValidatedDocument


class ExcelService:
    def generate(self, validated: ValidatedDocument, output_path: Path) -> Path:
        try:
            output_path.parent.mkdir(parents=True, exist_ok=True)
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                if validated.document_type == DocumentType.bank_statement:
                    self._write_bank_summary(writer, validated)
                    self._write_bank_statement(writer, validated)
                else:
                    self._write_summary(writer, validated)
                    self._write_line_items(writer, validated)
                self._write_validation_report(writer, validated)
                self._format_workbook(writer.book)
            return output_path
        except Exception as exc:
            raise ExcelGenerationFailureError(str(exc)) from exc

    def _write_summary(self, writer: pd.ExcelWriter, validated: ValidatedDocument) -> None:
        data = validated.data
        if validated.document_type == DocumentType.financial_report:
            summary = {
                "Report Name": data.get("report_name"),
                "Period": data.get("period"),
                "Currency": data.get("currency"),
            }
        else:
            summary = {
                "Document Number": data.get("invoice_number") or data.get("receipt_number") or data.get("po_number"),
                "Date": data.get("invoice_date") or data.get("date") or data.get("po_date"),
                "Vendor": data.get("vendor_name") or data.get("merchant_name") or data.get("buyer_name"),
                "Total": data.get("grand_total") or data.get("total"),
                "Tax": data.get("total_tax") or data.get("tax"),
                "Currency": data.get("currency"),
            }
        pd.DataFrame(summary.items(), columns=["Field", "Value"]).to_excel(writer, sheet_name="Document Summary", index=False)

    def _write_line_items(self, writer: pd.ExcelWriter, validated: ValidatedDocument) -> None:
        if validated.document_type == DocumentType.financial_report:
            metrics = validated.data.get("metrics") or {}
            if metrics:
                pd.DataFrame(metrics.items(), columns=["Metric", "Value"]).to_excel(writer, sheet_name="Report Metrics", index=False)
            tables = validated.data.get("tables") or []
            for idx, table in enumerate(tables):
                if isinstance(table, list):
                    df = pd.DataFrame(table)
                elif isinstance(table, dict):
                    df = pd.DataFrame([table])
                else:
                    continue
                df.to_excel(writer, sheet_name=f"Table {idx+1}", index=False)
        else:
            rows = validated.data.get("line_items") or []
            if rows:
                df = pd.DataFrame(rows).rename(
                    columns={"description": "Description", "quantity": "Qty", "unit_price": "Unit Price", "amount": "Amount"}
                )
            else:
                df = pd.DataFrame(columns=["Description", "Qty", "Unit Price", "Amount"])
            df.to_excel(writer, sheet_name="Line Items", index=False)

    def _write_bank_summary(self, writer: pd.ExcelWriter, validated: ValidatedDocument) -> None:
        data = validated.data
        rows = data.get("transactions") or []
        summary = {
            "Bank Name": data.get("bank_name"),
            "Account Number": data.get("account_number"),
            "Statement Period": data.get("statement_period"),
            "Total Debit": self._sum_column(rows, "debit"),
            "Total Credit": self._sum_column(rows, "credit"),
            "Closing Balance": rows[-1].get("balance") if rows else None,
        }
        pd.DataFrame(summary.items(), columns=["Field", "Value"]).to_excel(writer, sheet_name="Document Summary", index=False)

    def _write_bank_statement(self, writer: pd.ExcelWriter, validated: ValidatedDocument) -> None:
        rows = validated.data.get("transactions") or []
        df = pd.DataFrame(rows).rename(
            columns={"date": "Date", "description": "Description", "debit": "Debit", "credit": "Credit", "balance": "Balance"}
        )
        df.to_excel(writer, sheet_name="Transactions", index=False)
        sheet = writer.book["Transactions"]
        start = len(df) + 3
        sheet.cell(start, 1, "Total Debit")
        sheet.cell(start, 2, self._sum_column(rows, "debit"))
        sheet.cell(start + 1, 1, "Total Credit")
        sheet.cell(start + 1, 2, self._sum_column(rows, "credit"))
        sheet.cell(start + 2, 1, "Closing Balance")
        sheet.cell(start + 2, 2, rows[-1].get("balance") if rows else None)

        bold_font = Font(bold=True)
        for row_idx in range(start, start + 3):
            sheet.cell(row_idx, 1).font = bold_font
            sheet.cell(row_idx, 2).font = bold_font

    def _write_validation_report(self, writer: pd.ExcelWriter, validated: ValidatedDocument) -> None:
        rows = []
        issues_by_field: dict[str, list[Any]] = {}
        for issue in validated.issues:
            issues_by_field.setdefault(issue.field, []).append(issue)

        for field, score in validated.field_confidence.items():
            if field in ("line_items", "transactions", "tables", "metrics"):
                continue
            field_issues = issues_by_field.get(field, [])
            if field_issues:
                for issue in field_issues:
                    rows.append({
                        "Field": field,
                        "Confidence": score.confidence,
                        "Status": issue.status,
                        "Message": issue.message
                    })
            else:
                rows.append({
                    "Field": field,
                    "Confidence": score.confidence,
                    "Status": "ok" if score.confidence > 0.0 else "missing",
                    "Message": "" if score.confidence > 0.0 else f"Field '{field}' is missing"
                })

        for field in ("line_items", "transactions", "tables", "metrics"):
            field_issues = issues_by_field.get(field, [])
            for issue in field_issues:
                rows.append({
                    "Field": field,
                    "Confidence": issue.confidence,
                    "Status": issue.status,
                    "Message": issue.message
                })

        df = pd.DataFrame(rows, columns=["Field", "Confidence", "Status", "Message"])
        df.to_excel(writer, sheet_name="Validation Report", index=False)

    def _format_workbook(self, workbook: Any) -> None:
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for sheet in workbook.worksheets:
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            for column_cells in sheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 50)
            sheet.freeze_panes = "A2"

    def _sum_column(self, rows: list[dict[str, Any]], key: str) -> float:
        total = 0.0
        for row in rows:
            value = row.get(key) or 0
            try:
                total += float(value)
            except (TypeError, ValueError):
                continue
        return total

