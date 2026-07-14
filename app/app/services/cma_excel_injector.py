import logging
import openpyxl
from pathlib import Path
from typing import Optional, Dict, Any

from app.schemas.cma_excel_map import CMA_ROW_MAP, DEPRECIATION_ROW_MAP

logger = logging.getLogger(__name__)

def inject_cma_data_from_json(
    excel_path: Path, 
    json_data: dict, 
    target_year: str, 
    override_vals: Optional[Dict[int, Any]] = None
) -> Path:
    """
    Intelligently injects extracted CMA financial data from JSON into the master Excel sheet.
    
    Args:
        excel_path: Path to the master Excel file.
        json_data: Validated extraction JSON containing 'cma_data'.
        target_year: Target financial year as a string (e.g., '2024-25' or '2025').
        override_vals: Optional dictionary of Row -> Value overrides to apply manual adjustments.
        
    Returns:
        Path to the updated Excel file.
    """
    # Normalize target_year (e.g., '2024-25' -> '2025')
    year_short = target_year.split("-")[-1]
    if len(year_short) == 2:
        year_short = "20" + year_short  # 25 -> 2025
        
    logger.info(f"Starting Excel injection for FY {year_short} (source year: {target_year})")
    
    # Load workbook preserving formulas
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    
    # ─────────────────────────────────────────────────────────────────────────
    # 1. UPDATE 'CMA' SHEET
    # ─────────────────────────────────────────────────────────────────────────
    if "CMA" not in wb.sheetnames:
        raise ValueError("Worksheet 'CMA' not found in workbook.")
        
    ws_cma = wb["CMA"]
    
    # Find the target column in CMA sheet by searching Row 11 for the short year (e.g., '2025')
    target_col = None
    for col in range(1, ws_cma.max_column + 1):
        val = ws_cma.cell(row=11, column=col).value
        if val is not None and str(val).strip() == year_short:
            target_col = col
            break
            
    if target_col is None:
        raise ValueError(f"Could not find column for year '{year_short}' in CMA sheet Row 11.")
        
    col_letter = openpyxl.utils.get_column_letter(target_col)
    logger.info(f"Target year '{year_short}' mapped to Column {col_letter} (Index {target_col})")
    
    # ── Determine the correct number format from a populated reference column ──
    # Column S may have been empty/unused and defaults to format '0' (integer).
    # We must copy the number format from a known-good column (e.g. Column H = index 8)
    # to preserve decimal precision (e.g. '0.00') for all injected financial values.
    REF_COL = 8  # Column H — always populated with audited data
    ref_num_format = ws_cma.cell(row=14, column=REF_COL).number_format  # Row 14 is a numeric data row
    if not ref_num_format or ref_num_format == 'General':
        ref_num_format = '0.00'  # Fallback to standard 2-decimal financial format
    logger.info(f"Reference number format from Column H: '{ref_num_format}' — will apply to all cells in Column {col_letter}")
    
    # Inject mapped fields
    cma_data = json_data.get("cma_data", {})
    
    for row, config in CMA_ROW_MAP.items():
        cell = ws_cma.cell(row=row, column=target_col)
        
        # Copy number format from the same row in the reference column for consistency.
        # This ensures each row inherits the correct format (e.g. percentage rows vs. value rows).
        row_ref_fmt = ws_cma.cell(row=row, column=REF_COL).number_format
        if row_ref_fmt and row_ref_fmt != 'General':
            cell.number_format = row_ref_fmt
        else:
            cell.number_format = ref_num_format
        
        # Priority 1: User override values
        if override_vals and row in override_vals:
            cell.value = override_vals[row]
            logger.debug(f"Row {row:03d} (Override): Wrote '{cell.value}' fmt='{cell.number_format}'")
            continue
            
        # Priority 2: Configured Excel formulas (retaining/writing custom formulas)
        if config["type"] == "formula":
            # Adjust column reference in formula if needed (e.g. replace 'J' with col_letter)
            orig_formula = config["formula"]
            adjusted_formula = orig_formula.replace("J", col_letter).replace("I", openpyxl.utils.get_column_letter(target_col - 1))
            cell.value = adjusted_formula
            logger.debug(f"Row {row:03d} (Formula): Wrote '{cell.value}' fmt='{cell.number_format}'")
            continue
            
        # Priority 3: Retrieve value from JSON
        sec = config["section"]
        field = config["field"]
        
        # Look up key under both the short year ('2025') and full year ('2024-25')
        val = None
        field_data = cma_data.get(sec, {}).get("fields", {}).get(field, {})
        for y_key in [target_year, f"20{target_year}" if len(target_year)==2 else target_year, year_short, f"{int(year_short)-1}-{year_short[-2:]}"]:
            if y_key in field_data:
                val = field_data[y_key].get("value")
                if val is not None:
                    break
                    
        # Apply defaults if specified
        if val is None and "default" in config:
            val = config["default"]
            
        # Write to cell
        if val is not None:
            # Clean/parse values if numeric type
            if config["type"] == "number":
                try:
                    val = float(val)
                except (ValueError, TypeError):
                    pass
            cell.value = val
            logger.info(f"Row {row:03d} ({field}): Wrote '{val}' fmt='{cell.number_format}'")
        else:
            logger.warning(f"Row {row:03d} ({field}): Value not found in JSON for year {target_year}")

    # ─────────────────────────────────────────────────────────────────────────
    # 1b. APPLY OVERRIDES for rows NOT in CMA_ROW_MAP
    # ─────────────────────────────────────────────────────────────────────────
    # Some override rows (e.g. 138=General Reserve, 145=Drawings) are not in
    # CMA_ROW_MAP and were skipped during the main injection loop.
    if override_vals:
        for row, oval in override_vals.items():
            if row not in CMA_ROW_MAP:
                cell = ws_cma.cell(row=row, column=target_col)
                # Apply number format from reference column
                row_ref_fmt = ws_cma.cell(row=row, column=REF_COL).number_format
                if row_ref_fmt and row_ref_fmt != 'General':
                    cell.number_format = row_ref_fmt
                else:
                    cell.number_format = ref_num_format
                cell.value = oval
                logger.info(f"Row {row:03d} (Override-extra): Wrote '{oval}' fmt='{cell.number_format}'")

    # ─────────────────────────────────────────────────────────────────────────
    # 1c. COPY FORMULAS & NUMBER FORMATS for rows not in CMA_ROW_MAP
    # ─────────────────────────────────────────────────────────────────────────
    # Many rows (subtotals, totals, computed fields) are formula-only and are
    # not in CMA_ROW_MAP. We copy their formulas from the reference column (H)
    # to the target column, adjusting column letter references. We also ensure
    # every cell in the data range inherits the correct number format.
    ref_col_letter = openpyxl.utils.get_column_letter(REF_COL)
    handled_rows = set(CMA_ROW_MAP.keys())
    if override_vals:
        handled_rows.update(override_vals.keys())
    
    import re as _re
    DATA_ROW_START = 12
    DATA_ROW_END = 210
    formula_copy_count = 0
    format_copy_count = 0
    
    for row in range(DATA_ROW_START, DATA_ROW_END + 1):
        target_cell = ws_cma.cell(row=row, column=target_col)
        ref_cell = ws_cma.cell(row=row, column=REF_COL)
        
        # Always copy the number format from the reference column
        if ref_cell.number_format and ref_cell.number_format != 'General':
            if target_cell.number_format != ref_cell.number_format:
                target_cell.number_format = ref_cell.number_format
                format_copy_count += 1
        
        # For rows not in the map and not overridden, copy formulas from reference column
        if row not in handled_rows and target_cell.value is None:
            ref_val = ref_cell.value
            if isinstance(ref_val, str) and ref_val.startswith("="):
                # Replace reference column letter with target column letter
                # Use word-boundary-aware replacement to avoid mangling cell refs like "SH" 
                adjusted = _re.sub(
                    rf'\b{ref_col_letter}(\d+)\b',
                    lambda m: f'{col_letter}{m.group(1)}',
                    ref_val
                )
                target_cell.value = adjusted
                formula_copy_count += 1
                logger.debug(f"Row {row:03d} (FormulaCopy): '{ref_val}' → '{adjusted}' fmt='{target_cell.number_format}'")
    
    logger.info(f"Formula copy pass: {formula_copy_count} formulas copied, {format_copy_count} formats updated (rows {DATA_ROW_START}-{DATA_ROW_END})")

    # ─────────────────────────────────────────────────────────────────────────
    # 2. UPDATE 'DEPRECIATION' SHEET
    # ─────────────────────────────────────────────────────────────────────────
    if "DEPRECIATION" in wb.sheetnames:
        ws_dep = wb["DEPRECIATION"]
        logger.info("Updating 'DEPRECIATION' sheet opening balances...")
        
        # Find opening columns for Gross Block & Depreciation by searching Row 3 for the date
        # date format example: '01.04.2025'
        date_query = f"01.04.{year_short}"
        dep_gross_col = None
        
        for col in range(1, ws_dep.max_column + 1):
            val = ws_dep.cell(row=3, column=col).value
            if val is not None and date_query in str(val):
                dep_gross_col = col
                break
                
        if dep_gross_col is not None:
            dep_accum_col = dep_gross_col + 3
            logger.info(f"Depreciation opening date '{date_query}' mapped to Col {openpyxl.utils.get_column_letter(dep_gross_col)} (Gross Block) and Col {openpyxl.utils.get_column_letter(dep_accum_col)} (Accum. Depreciation)")
            
            # Since individual asset classes opening values can be extracted or manual, we let it fall back gracefully
            # For our current template, Note 11 asset details are written
            # We fetch them from the fixed_assets section or allow override
            for row, dep_config in DEPRECIATION_ROW_MAP.items():
                asset_name = dep_config["asset"]
                
                # Check overrides first
                if override_vals and f"dep_gross_{row}" in override_vals:
                    ws_dep.cell(row=row, column=dep_gross_col).value = override_vals[f"dep_gross_{row}"]
                if dep_config["has_dep"] and override_vals and f"dep_accum_{row}" in override_vals:
                    ws_dep.cell(row=row, column=dep_accum_col).value = override_vals[f"dep_accum_{row}"]
        else:
            logger.warning(f"Could not find opening date '{date_query}' in DEPRECIATION sheet Row 3.")

    # Save the updated workbook
    wb.save(excel_path)
    logger.info(f"Successfully saved updated Excel workbook at: {excel_path}")
    return excel_path
